import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

print("=== SIMPLE FIX: CONSOLIDATE WEST VARIANTS ===\n")

# Read the overall sheet
df = pd.read_excel('data_exports/Depot_OnTime_Performance.xlsx', sheet_name='Overall On-Time Performance')

print("Before fix:")
print(df.to_string(index=False))

# Get WEST variants (case-insensitive, strip whitespace)
west_mask = df['Depot'].str.strip().str.upper() == 'WEST'
west_rows = df[west_mask]

print(f"\n✓ Found {len(west_rows)} WEST variant rows")

# Sum all WEST variants
x_total_deliveries = west_rows['Total Deliveries'].sum()
x_ontime = west_rows['On-Time'].sum()
x_late = west_rows['Late'].sum()
x_ontime_pct = (x_ontime / x_total_deliveries * 100).round(1)

print(f"✓ Summed values:")
print(f"    Total Deliveries: {x_total_deliveries}")
print(f"    On-Time: {x_ontime}")
print(f"    Late: {x_late}")
print(f"    On-Time %: {x_ontime_pct}")

# Remove all WEST variants
df = df[~west_mask]

# Add consolidated WEST row
df = pd.concat([df, pd.DataFrame({
    'Depot': ['WEST'],
    'Total Deliveries': [x_total_deliveries],
    'On-Time': [x_ontime],
    'Late': [x_late],
    'On-Time %': [x_ontime_pct]
})], ignore_index=True)

# Sort by On-Time %
df = df.sort_values('On-Time %', ascending=False).reset_index(drop=True)

print(f"\nAfter fix:")
print(df.to_string(index=False))

# Write back to Excel
wb = load_workbook('data_exports/Depot_OnTime_Performance.xlsx')
ws = wb['Overall On-Time Performance']

# Clear and rewrite
ws.delete_rows(2, ws.max_row)

for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
    for c_idx, value in enumerate(row, start=1):
        ws.cell(row=r_idx, column=c_idx, value=value)

# Format
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

for cell in ws[1]:
    if cell.value:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

for row_idx in range(2, ws.max_row + 1):
    ontime_pct = ws.cell(row=row_idx, column=5).value
    for col_idx in range(1, 6):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if col_idx == 5:
            if ontime_pct and ontime_pct >= 90:
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif ontime_pct and ontime_pct >= 80:
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

wb.save('data_exports/Depot_OnTime_Performance.xlsx')
print("\n✅ Fixed! Saved to Depot_OnTime_Performance.xlsx")
