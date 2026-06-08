import pandas as pd
import sqlite3
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

print("📊 Loading Depot Monthly Breakdown from Monthly_Feedback_Enhanced.xlsx...")

# Read the existing Depot Monthly Breakdown sheet
existing_df = pd.read_excel('data_exports/Monthly_Feedback_Enhanced.xlsx', sheet_name='Depot Monthly Breakdown')
print(f"✅ Loaded existing data: {existing_df.shape}")
print("\nExisting columns:")
print(existing_df.columns.tolist())
print("\nFirst few rows:")
print(existing_df.head())

# Load on-time percentage data from database
print("\n📊 Loading on-time percentage data from database...")
conn = sqlite3.connect('delivery.db')

query = """
SELECT 
    strftime('%Y-%m', d.booking_datetime) as Month,
    CASE 
        WHEN UPPER(TRIM(d.branch)) IN ('CENTRAL', 'CNETRAL') THEN 'Central'
        WHEN UPPER(TRIM(d.branch)) = 'EAST' THEN 'East'
        WHEN UPPER(TRIM(d.branch)) IN ('NORTH', 'NOTH') THEN 'North'
        WHEN UPPER(TRIM(d.branch)) = 'WEST' THEN 'West'
        ELSE d.branch
    END as Depot,
    ROUND(AVG(CASE WHEN d.delivery_datetime <= d.promised_delivery_datetime THEN 1 ELSE 0 END) * 100, 2) as OnTime_Percentage
FROM deliveries d
GROUP BY Month, Depot
ORDER BY Month, Depot
"""

ontime_df = pd.read_sql_query(query, conn)
conn.close()

print(f"✅ Loaded on-time data: {ontime_df.shape}")
print(ontime_df.head(10))

# Merge with existing data
print("\n🔗 Merging datasets...")

# Identify how the existing data is structured
if 'Month' in existing_df.columns and 'Depot' in existing_df.columns:
    merged_df = existing_df.merge(ontime_df, on=['Month', 'Depot'], how='left')
    print("✅ Merged on [Month, Depot]")
elif 'month' in existing_df.columns.str.lower():
    # Handle case variations
    existing_df_renamed = existing_df.copy()
    for col in existing_df_renamed.columns:
        if col.lower() == 'month':
            existing_df_renamed.rename(columns={col: 'Month'}, inplace=True)
        elif col.lower() == 'depot':
            existing_df_renamed.rename(columns={col: 'Depot'}, inplace=True)
    merged_df = existing_df_renamed.merge(ontime_df, on=['Month', 'Depot'], how='left')
    print("✅ Merged on [Month, Depot] (case-insensitive)")
else:
    print("⚠️  Could not find Month/Depot columns, checking structure...")
    print(existing_df.head(15))
    merged_df = existing_df.copy()

# Rename the new column
merged_df.rename(columns={'OnTime_Percentage': 'OnTime_%'}, inplace=True)

print("\n" + "="*140)
print("📊 UPDATED DEPOT MONTHLY BREAKDOWN WITH ON-TIME %")
print("="*140 + "\n")
print(merged_df.to_string(index=False))

# Save as CSV
csv_path = 'data_exports/Depot_Monthly_Breakdown_With_OnTime.csv'
merged_df.to_csv(csv_path, index=False)
print(f"\n✅ Saved CSV: {csv_path}")

# Create Excel file with formatting
print("\n📁 Creating formatted Excel file...")

wb = load_workbook('data_exports/Monthly_Feedback_Enhanced.xlsx')

# Check if sheet exists, if so remove it
if 'Depot Monthly Breakdown Enhanced' in wb.sheetnames:
    del wb['Depot Monthly Breakdown Enhanced']

# Create new sheet
ws = wb.create_sheet('Depot Monthly Breakdown Enhanced', index=1)

# Formatting
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                     top=Side(style='thin'), bottom=Side(style='thin'))
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)

# Add title
ws['A1'] = "Depot Monthly Breakdown with On-Time Performance"
ws['A1'].font = Font(bold=True, size=12)

# Add headers and data
start_row = 3
for col_idx, col_name in enumerate(merged_df.columns, 1):
    cell = ws.cell(row=start_row, column=col_idx, value=col_name)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Add data rows
for row_idx, row_data in enumerate(merged_df.values, start_row + 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Format numbers
        if isinstance(value, float):
            if 'OnTime_%' in merged_df.columns[col_idx - 1] or '%' in str(merged_df.columns[col_idx - 1]):
                cell.number_format = '0.00'
            else:
                cell.number_format = '0.00'

# Auto-fit columns
for col_idx in range(1, len(merged_df.columns) + 1):
    max_length = 0
    col_letter = get_column_letter(col_idx)
    for row in ws[col_letter]:
        try:
            if len(str(row.value)) > max_length:
                max_length = len(str(row.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 40)
    ws.column_dimensions[col_letter].width = adjusted_width

# Freeze header
ws.freeze_panes = f'A{start_row + 1}'

# Save
excel_path = 'data_exports/Monthly_Feedback_Enhanced_Updated.xlsx'
wb.save(excel_path)
print(f"✅ Saved Excel: {excel_path}")

# Summary
print("\n" + "="*140)
print("🎯 KEY METRICS")
print("="*140)

print("\nOn-Time % by Depot (Average across all months):")
depot_avg = merged_df.groupby('Depot')['OnTime_%'].mean().round(2).sort_values(ascending=False)
for depot, pct in depot_avg.items():
    print(f"  {depot}: {pct}%")

print("\nOn-Time % by Month (Average across all depots):")
month_avg = merged_df.groupby('Month')['OnTime_%'].mean().round(2).sort_values(ascending=False)
for month, pct in month_avg.items():
    print(f"  {month}: {pct}%")

print("\n" + "="*140)
