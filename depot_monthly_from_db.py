import sqlite3
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

print("📊 Loading data from delivery.db...")

# Connect to database
conn = sqlite3.connect('delivery.db')

# Query deliveries with feedback
query = """
SELECT 
    d.delivery_id,
    d.branch,
    d.promised_delivery_datetime,
    d.delivery_datetime,
    CASE WHEN d.delivery_datetime <= d.promised_delivery_datetime THEN 1 ELSE 0 END as on_time,
    strftime('%Y-%m', d.booking_datetime) as booking_month
FROM deliveries d
"""

df = pd.read_sql_query(query, conn)
print(f"✅ Loaded: {len(df):,} rows")

# Standardize branch names
print("🏢 Standardizing branch names...")
df['branch_clean'] = df['branch'].str.strip().str.upper()
branch_mapping = {
    'CENTRAL': 'Central',
    'CNETRAL': 'Central',
    'EAST': 'East',
    'WEST': 'West',
    'NORTH': 'North',
    'NOTH': 'North'
}
df['branch_clean'] = df['branch_clean'].map(lambda x: branch_mapping.get(x, x))

# Create depot-month breakdown with on-time proportions
print("\n📈 Calculating depot monthly on-time proportions...")
depot_monthly = df.groupby(['booking_month', 'branch_clean']).agg({
    'delivery_id': 'count',
    'on_time': lambda x: (x.sum() / len(x) * 100).round(1)
}).reset_index()

depot_monthly.columns = ['Month', 'Depot', 'Total Deliveries', 'On-Time %']
depot_monthly['Total Deliveries'] = depot_monthly['Total Deliveries'].astype(int)

# Create pivot tables
pivot_depot_ontime = depot_monthly.pivot(index='Month', columns='Depot', values='On-Time %').round(1)
pivot_depot_counts = depot_monthly.pivot(index='Month', columns='Depot', values='Total Deliveries')

print("\n" + "="*100)
print("📊 DEPOT MONTHLY ON-TIME DELIVERY PROPORTION (%)")
print("="*100 + "\n")
print(pivot_depot_ontime.to_string())

print("\n" + "="*100)
print("📦 TOTAL DELIVERIES BY DEPOT & MONTH")
print("="*100 + "\n")
print(pivot_depot_counts.to_string())

# Export to Excel with formatting
print("\n📁 Creating Excel file...")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)

wb = Workbook()

# Sheet 1: On-Time Proportion
ws1 = wb.active
ws1.title = "On-Time %"
ws1['A1'] = "Depot Monthly On-Time Delivery Proportion (%)"
ws1['A1'].font = Font(bold=True, size=12)

start_row = 3
for r_idx, row in enumerate(dataframe_to_rows(pivot_depot_ontime.reset_index(), index=False, header=True), start_row):
    for c_idx, value in enumerate(row, 1):
        cell = ws1.cell(row=r_idx, column=c_idx, value=value)
        cell.border = thin_border
        if r_idx == start_row:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        else:
            if c_idx == 1:
                cell.alignment = Alignment(horizontal='left')
            else:
                cell.alignment = Alignment(horizontal='center')
                cell.number_format = '0.0'

for col_idx, col in enumerate(ws1.columns, 1):
    max_length = max(len(str(cell.value)) for cell in col if cell.value)
    ws1.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 20)

ws1.freeze_panes = f'A{start_row + 1}'

# Sheet 2: Delivery Counts
ws2 = wb.create_sheet("Total Deliveries")
ws2['A1'] = "Depot Monthly Total Deliveries"
ws2['A1'].font = Font(bold=True, size=12)

start_row = 3
for r_idx, row in enumerate(dataframe_to_rows(pivot_depot_counts.reset_index(), index=False, header=True), start_row):
    for c_idx, value in enumerate(row, 1):
        cell = ws2.cell(row=r_idx, column=c_idx, value=value)
        cell.border = thin_border
        if r_idx == start_row:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        else:
            if c_idx == 1:
                cell.alignment = Alignment(horizontal='left')
            else:
                cell.alignment = Alignment(horizontal='center')
                cell.number_format = '0'

for col_idx, col in enumerate(ws2.columns, 1):
    max_length = max(len(str(cell.value)) for cell in col if cell.value)
    ws2.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 20)

ws2.freeze_panes = f'A{start_row + 1}'

# Sheet 3: Detailed breakdown
ws3 = wb.create_sheet("Detailed")
ws3['A1'] = "Depot Monthly Breakdown (Detailed)"
ws3['A1'].font = Font(bold=True, size=12)

start_row = 3
for r_idx, row in enumerate(dataframe_to_rows(depot_monthly, index=False, header=True), start_row):
    for c_idx, value in enumerate(row, 1):
        cell = ws3.cell(row=r_idx, column=c_idx, value=value)
        cell.border = thin_border
        if r_idx == start_row:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        else:
            cell.alignment = Alignment(horizontal='center')
            if c_idx == 4:
                cell.number_format = '0.0'

for col_idx, col in enumerate(ws3.columns, 1):
    max_length = max(len(str(cell.value)) for cell in col if cell.value)
    ws3.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 25)

ws3.freeze_panes = f'A{start_row + 1}'

wb.save('data_exports/Depot_Monthly_OnTime_Proportion.xlsx')
print(f"✅ Saved: data_exports/Depot_Monthly_OnTime_Proportion.xlsx\n")

# Summary stats
print("="*100)
print("🎯 KEY INSIGHTS")
print("="*100)

best_idx = depot_monthly['On-Time %'].idxmax()
best = depot_monthly.iloc[best_idx]
print(f"\nBest Performing Depot-Month:")
print(f"  → {best['Depot']} in {best['Month']}: {best['On-Time %']:.1f}% on-time ({int(best['Total Deliveries'])} deliveries)")

worst_idx = depot_monthly['On-Time %'].idxmin()
worst = depot_monthly.iloc[worst_idx]
print(f"\nWorst Performing Depot-Month:")
print(f"  → {worst['Depot']} in {worst['Month']}: {worst['On-Time %']:.1f}% on-time ({int(worst['Total Deliveries'])} deliveries)")

print(f"\nDepot Overall Performance (All Months):")
depot_avg = df.groupby('branch_clean').agg({
    'on_time': lambda x: (x.sum() / len(x) * 100).round(1),
    'delivery_id': 'count'
}).sort_values('on_time', ascending=False)
depot_avg.columns = ['On-Time %', 'Total Deliveries']
print(depot_avg.to_string())

print(f"\nMonthly Trend (Overall):")
monthly_trend = df.groupby('booking_month')['on_time'].apply(lambda x: (x.sum() / len(x) * 100)).reset_index()
monthly_trend.columns = ['Month', 'On-Time %']
for _, row in monthly_trend.iterrows():
    print(f"  {row['Month']}: {row['On-Time %']:.1f}%")

conn.close()
print("\n" + "="*100)
