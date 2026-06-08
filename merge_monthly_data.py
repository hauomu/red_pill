import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import os

print("📊 Loading data files...")

# Load Monthly Feedback Enhanced (from existing file or recreate)
try:
    # Try loading from Excel first
    feedback_df = pd.read_excel('data_exports/Monthly_Feedback_Enhanced.xlsx', sheet_name='Monthly Summary')
    print("✅ Loaded Monthly_Feedback_Enhanced.xlsx")
except:
    print("⚠️  Monthly_Feedback_Enhanced.xlsx not found, creating from database...")
    import sqlite3
    conn = sqlite3.connect('delivery.db')
    
    # Query monthly feedback metrics
    query = """
    SELECT 
        strftime('%Y-%m', f.feedback_datetime) as booking_month,
        COUNT(DISTINCT f.feedback_id) as total_feedback,
        ROUND(AVG(f.rating), 2) as avg_rating,
        MIN(f.rating) as min_rating,
        MAX(f.rating) as max_rating,
        ROUND(STDEV(f.rating), 2) as std_rating
    FROM feedback f
    GROUP BY booking_month
    ORDER BY booking_month
    """
    feedback_df = pd.read_sql_query(query, conn)
    feedback_df['booking_month'] = feedback_df['booking_month'].astype(str)
    conn.close()
    print("✅ Created from database")

# Load Depot Monthly On-Time Proportion
depot_ontime_df = pd.read_excel('data_exports/Depot_Monthly_OnTime_Proportion.xlsx', sheet_name='On-Time %')
print("✅ Loaded Depot_Monthly_OnTime_Proportion.xlsx")

# Reset index to make Month a column
if 'Month' in depot_ontime_df.columns:
    pass
else:
    depot_ontime_df = depot_ontime_df.reset_index()

print("\n📋 Data Preview:")
print("\nFeedback data:")
print(feedback_df.head())
print("\nDepot On-Time data:")
print(depot_ontime_df.head())

# Standardize month columns
print("\n🔗 Merging datasets...")
feedback_df.rename(columns={'booking_month': 'Month'}, inplace=True)
if 'Month' not in feedback_df.columns:
    # If no Month column, take first column as Month
    feedback_df.rename(columns={feedback_df.columns[0]: 'Month'}, inplace=True)

if 'Month' not in depot_ontime_df.columns:
    depot_ontime_df.rename(columns={depot_ontime_df.columns[0]: 'Month'}, inplace=True)

# Convert Month to string for matching
feedback_df['Month'] = feedback_df['Month'].astype(str)
depot_ontime_df['Month'] = depot_ontime_df['Month'].astype(str)

print(f"\nFeedback months: {feedback_df['Month'].tolist()}")
print(f"Depot months: {depot_ontime_df['Month'].tolist()}")

# Merge on Month
merged_df = feedback_df.merge(depot_ontime_df, on='Month', how='outer')
merged_df = merged_df.sort_values('Month')

print("\n" + "="*120)
print("📊 MERGED: MONTHLY FEEDBACK + DEPOT ON-TIME PROPORTION")
print("="*120 + "\n")
print(merged_df.to_string(index=False))

# Save as CSV
csv_path = 'data_exports/Monthly_Feedback_Depot_OnTime_Merged.csv'
merged_df.to_csv(csv_path, index=False)
print(f"\n✅ Saved CSV: {csv_path}")

# Save as Excel with formatting
print("📁 Creating Excel file...")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)

wb = Workbook()
ws = wb.active
ws.title = "Monthly Summary"

# Add title
ws['A1'] = "Monthly Feedback & Depot On-Time Performance"
ws['A1'].font = Font(bold=True, size=12)

# Add data starting from row 3
start_row = 3
for r_idx, row in enumerate(dataframe_to_rows(merged_df, index=False, header=True), start_row):
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        cell.border = thin_border
        if r_idx == start_row:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        else:
            if c_idx == 1:  # Month column
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.alignment = Alignment(horizontal='center')
                # Format numbers
                if isinstance(value, (int, float)):
                    if 'rating' in merged_df.columns[c_idx - 1].lower() or '%' in str(value):
                        cell.number_format = '0.0'
                    else:
                        cell.number_format = '0'

# Auto-fit columns
for col_idx, col in enumerate(ws.columns, 1):
    max_length = max(len(str(cell.value)) for cell in col if cell.value)
    ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 35)

ws.freeze_panes = f'A{start_row + 1}'

excel_path = 'data_exports/Monthly_Feedback_Depot_OnTime_Merged.xlsx'
wb.save(excel_path)
print(f"✅ Saved Excel: {excel_path}\n")

# Summary statistics
print("="*120)
print("🎯 KEY INSIGHTS - CORRELATION ANALYSIS")
print("="*120)

# Find correlations between feedback rating and on-time %
if 'avg_rating' in merged_df.columns and 'Central' in merged_df.columns:
    corr = merged_df[['avg_rating', 'Central', 'East', 'North', 'West']].corr()
    print("\nCorrelation: Average Rating vs Depot On-Time %")
    print(corr)

print("\nTrend Analysis:")
print("Month | Avg Rating | Central % | East % | North % | West %")
for _, row in merged_df.iterrows():
    month = row['Month']
    rating = row.get('avg_rating', '-')
    central = row.get('Central', '-')
    east = row.get('East', '-')
    north = row.get('North', '-')
    west = row.get('West', '-')
    print(f"{month} | {rating} | {central} | {east} | {north} | {west}")

print("\n" + "="*120)
