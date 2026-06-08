import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

print("Generating monthly feedback scores...\n")

df = pd.read_csv('data_exports/deliveries_feedback_merged.csv', encoding='latin-1')
print(f"✅ Loaded: {len(df):,} rows")

# Parse dates
df['booking_dt'] = pd.to_datetime(df['booking_datetime'], format='%d/%m/%Y %H:%M', errors='coerce')
df['month'] = df['booking_dt'].dt.to_period('M').astype(str)

# Monthly summary
monthly = df.groupby('month').agg({
    'delivery_id': 'count',
    'distance_km': 'mean',
    'driver_experience_months': 'mean'
}).round(2)

# On-time rate
df['on_time'] = (df['delivery_datetime'] <= df['promised_delivery_datetime']).astype(int)
ontime = df.groupby('month')['on_time'].apply(lambda x: (x.sum() / len(x) * 100).round(1))

# Feedback scores
feedback_df = df[df['rating'].notna()]
feedback = feedback_df.groupby('month')['rating'].agg(['count', 'mean', 'min', 'max', 'std']).round(2)

# Combine
result = pd.DataFrame({
    'Month': monthly.index,
    'Total Deliveries': monthly['delivery_id'].values,
    'Total Feedback': feedback['count'].values if len(feedback) > 0 else 0,
    'Avg Rating': feedback['mean'].values if len(feedback) > 0 else None,
    'Min Rating': feedback['min'].values if len(feedback) > 0 else None,
    'Max Rating': feedback['max'].values if len(feedback) > 0 else None,
    'Std Dev': feedback['std'].values if len(feedback) > 0 else None,
    'Avg Distance (km)': monthly['distance_km'].values,
    'Avg Driver Exp (mo)': monthly['driver_experience_months'].values,
    'On-time %': ontime.values if len(ontime) > 0 else None
})

# Excel formatting
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)

wb = Workbook()
ws = wb.active
ws.title = "Monthly Summary"

for r_idx, row in enumerate(dataframe_to_rows(result, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        cell.border = thin_border
        if r_idx == 1:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        else:
            cell.alignment = Alignment(horizontal='center', vertical='center')

for col in ws.columns:
    max_len = max(len(str(cell.value)) for cell in col if cell.value)
    ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 50)

ws.freeze_panes = 'A2'
wb.save('data_exports/Monthly_Feedback_Scores.xlsx')
print(f"\n✅ File saved: data_exports/Monthly_Feedback_Scores.xlsx\n")

print("="*100)
print(result.to_string(index=False))
print("="*100)
