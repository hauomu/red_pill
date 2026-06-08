import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

print("Creating Deliveries_Main file with borders and formatting...\n")

# Read the merged CSV file
merged_path = 'data_exports/deliveries_feedback_merged.csv'
print(f"📖 Reading merged file: {merged_path}")
df_merged = pd.read_csv(merged_path, encoding='latin-1')
print(f"   ✅ Loaded: {len(df_merged):,} rows x {len(df_merged.columns)} columns")

# Create Excel file
output_path = 'data_exports/Deliveries_Main.xlsx'
print(f"\n📝 Creating Excel file: {output_path}")

# Write to Excel
df_merged.to_excel(output_path, sheet_name='Deliveries', index=False)
print(f"   ✅ Initial file created")

# Load and format the Excel file
print(f"\n✨ Applying formatting...")
wb = load_workbook(output_path)
ws = wb.active

# Define border style
thin_border = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

# Header formatting
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)

# Apply borders and formatting to all cells
print(f"   Formatting {len(df_merged):,} data rows...")
for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=len(df_merged) + 1, 
                                           min_col=1, max_col=len(df_merged.columns)), 1):
    for cell in row:
        # Apply border to all cells
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
        
        # Format header row
        if row_idx == 1:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Auto-adjust column widths
print(f"   Adjusting column widths...")
for col_idx, col in enumerate(ws.columns, 1):
    max_length = 0
    column_letter = get_column_letter(col_idx)
    
    for cell in col:
        try:
            if cell.value:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        except:
            pass
    
    adjusted_width = min(max_length + 2, 50)
    ws.column_dimensions[column_letter].width = adjusted_width

# Freeze header row
ws.freeze_panes = 'A2'

# Set row height for header
ws.row_dimensions[1].height = 25

# Save formatted Excel file
wb.save(output_path)
print(f"   ✅ Formatting applied")

print(f"\n" + "="*70)
print(f"✅ DELIVERIES_MAIN FILE CREATED SUCCESSFULLY!")
print(f"="*70)

summary = f"""
📊 FILE DETAILS:
   Name: Deliveries_Main.xlsx
   Location: data_exports/
   Total Rows: {len(df_merged):,}
   Total Columns: {len(df_merged.columns)}
   File Size: {os.path.getsize(output_path) / 1024 / 1024:.2f} MB

✨ FORMATTING APPLIED:
   ✓ All cells have row & column borders
   ✓ Header row: Blue background with white text
   ✓ Auto-fitted column widths
   ✓ Center-aligned header
   ✓ Left-aligned data
   ✓ Frozen header row for scrolling
   ✓ 25px header height

📋 COLUMNS ({len(df_merged.columns)} total):
"""
for i, col in enumerate(df_merged.columns, 1):
    summary += f"   {i:2}. {col}\n"

summary += f"""
📂 READY TO USE:
   • Open in Microsoft Excel
   • Open in Google Sheets
   • Open in LibreOffice Calc
   
🚀 Next Steps:
   1. Open Deliveries_Main.xlsx
   2. Use for Exploratory Data Analysis (EDA)
   3. Use for Machine Learning Pipeline
   4. Export specific columns as needed
"""

print(summary)

print(f"\n✅ File saved successfully at: {output_path}")
