import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import re

print("Generating enhanced monthly feedback analysis...\n")

df = pd.read_csv('data_exports/deliveries_feedback_merged.csv', encoding='latin-1')
print(f"✅ Loaded: {len(df):,} rows")

# Parse dates
df['booking_dt'] = pd.to_datetime(df['booking_datetime'], format='%d/%m/%Y %H:%M', errors='coerce')
df['month'] = df['booking_dt'].dt.to_period('M').astype(str)

# Filter feedback records
feedback_df = df[df['rating'].notna()].copy()
print(f"✅ Feedback records: {len(feedback_df):,}\n")

# Categorize sentiment
feedback_df['sentiment'] = feedback_df['rating'].apply(
    lambda x: 'Positive' if x >= 4 else ('Negative' if x <= 2 else 'Neutral')
)

# Complaint categories (regex patterns)
complaint_patterns = {
    'Late Delivery': r'(late|slow|hours late|promised window|delay)',
    'Poor Communication': r'(communication|no (update|notification)|not informed)',
    'Driver Behavior': r'(rude|unprofessional|careless|rough)',
    'Package Damage': r'(damage|broken|damaged|roughly handled)',
    'Other': r'.*'
}

def categorize_complaint(comment):
    """Categorize complaint based on keywords in comment"""
    if pd.isna(comment):
        return 'No Comment'
    comment_lower = str(comment).lower()
    for category, pattern in complaint_patterns.items():
        if re.search(pattern, comment_lower):
            return category
    return 'Other'

# Apply complaint categorization to negative feedback
negative_feedback = feedback_df[feedback_df['sentiment'] == 'Negative'].copy()
negative_feedback['complaint_category'] = negative_feedback['comment'].apply(categorize_complaint)

print("=" * 140)
print("COMPLAINT CATEGORIES DISTRIBUTION")
print("=" * 140)
complaint_dist = negative_feedback['complaint_category'].value_counts()
for cat, count in complaint_dist.items():
    pct = (count / len(negative_feedback) * 100)
    print(f"  {cat}: {count:,} ({pct:.1f}%)")

print("\n" + "=" * 140)
print("VEHICLE TYPE CORRELATION WITH COMPLAINTS")
print("=" * 140)
vehicle_complaint = pd.crosstab(
    negative_feedback['vehicle_type'], 
    negative_feedback['complaint_category'], 
    margins=True
)
print(vehicle_complaint)

print("\n" + "=" * 140)
print("MONTHLY ANALYSIS BY PRIORITY LEVEL")
print("=" * 140)

monthly_priority_data = []

for month in sorted(feedback_df['month'].unique()):
    month_data = feedback_df[feedback_df['month'] == month]
    
    print(f"\n📅 {month}:")
    for priority in ['VIP', 'Premium', 'Express', 'Standard']:
        priority_data = month_data[month_data['delivery_priority'] == priority]
        if len(priority_data) > 0:
            avg_rating = priority_data['rating'].mean()
            neg_count = len(priority_data[priority_data['sentiment'] == 'Negative'])
            neg_pct = (neg_count / len(priority_data) * 100)
            
            print(f"  {priority:10} - Avg Rating: {avg_rating:.2f} | Negative: {neg_count:3} ({neg_pct:.1f}%)")
            
            monthly_priority_data.append({
                'Month': month,
                'Priority': priority,
                'Total': len(priority_data),
                'Avg Rating': round(avg_rating, 2),
                'Negative': neg_count,
                'Negative %': round(neg_pct, 1)
            })

priority_result = pd.DataFrame(monthly_priority_data)

# Create comprehensive Excel with multiple sheets
wb = Workbook()

# Sheet 1: Priority Analysis
ws1 = wb.active
ws1.title = "Priority Breakdown"

thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=10)
negative_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

for c_idx, col_name in enumerate(priority_result.columns, 1):
    cell = ws1.cell(row=1, column=c_idx, value=col_name)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

for r_idx, (_, row) in enumerate(priority_result.iterrows(), 2):
    for c_idx, (col_name, value) in enumerate(row.items(), 1):
        cell = ws1.cell(row=r_idx, column=c_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if col_name in ['Negative', 'Negative %']:
            cell.fill = negative_fill

for col in ws1.columns:
    max_len = max(len(str(cell.value)) for cell in col if cell.value)
    ws1.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 20)

ws1.freeze_panes = 'A2'

# Sheet 2: Complaint Categories
ws2 = wb.create_sheet("Complaint Categories")

complaint_data = []
for cat in complaint_dist.index:
    count = complaint_dist[cat]
    pct = round(count / len(negative_feedback) * 100, 1)
    complaint_data.append({'Category': cat, 'Count': count, 'Percentage': pct})

complaint_df = pd.DataFrame(complaint_data)

for c_idx, col_name in enumerate(complaint_df.columns, 1):
    cell = ws2.cell(row=1, column=c_idx, value=col_name)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border

for r_idx, (_, row) in enumerate(complaint_df.iterrows(), 2):
    for c_idx, (col_name, value) in enumerate(row.items(), 1):
        cell = ws2.cell(row=r_idx, column=c_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

for col in ws2.columns:
    max_len = max(len(str(cell.value)) for cell in col if cell.value)
    ws2.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 20)

ws2.freeze_panes = 'A2'

# Sheet 3: Vehicle Complaint Correlation
ws3 = wb.create_sheet("Vehicle Correlation")

row_idx = 1
for col_idx, col_name in enumerate(vehicle_complaint.columns, 1):
    cell = ws3.cell(row=row_idx, column=col_idx, value=col_name)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border

for row_idx, (idx, row_data) in enumerate(vehicle_complaint.iterrows(), 2):
    ws3.cell(row=row_idx, column=1, value=idx).border = thin_border
    for col_idx, value in enumerate(row_data, 2):
        cell = ws3.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

for col in ws3.columns:
    max_len = max(len(str(cell.value)) for cell in col if cell.value)
    ws3.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 20)

wb.save('data_exports/Monthly_Feedback_Enhanced.xlsx')
print(f"\n✅ File saved: data_exports/Monthly_Feedback_Enhanced.xlsx")
print("\n📊 Sheets created:")
print("   1. Priority Breakdown - Ratings by priority level per month")
print("   2. Complaint Categories - Distribution of complaint types")
print("   3. Vehicle Correlation - Complaints by vehicle type")
