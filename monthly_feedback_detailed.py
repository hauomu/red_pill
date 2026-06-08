import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

print("Generating monthly feedback summary with sentiment-separated comments...\n")

df = pd.read_csv('data_exports/deliveries_feedback_merged.csv', encoding='latin-1')
print(f"✅ Loaded: {len(df):,} rows")

# Parse dates
df['booking_dt'] = pd.to_datetime(df['booking_datetime'], format='%d/%m/%Y %H:%M', errors='coerce')
df['month'] = df['booking_dt'].dt.to_period('M').astype(str)

# Filter feedback records only
feedback_df = df[df['rating'].notna()].copy()
print(f"✅ Feedback records: {len(feedback_df):,}")

# Categorize sentiment
feedback_df['sentiment'] = feedback_df['rating'].apply(
    lambda x: 'Positive' if x >= 4 else ('Negative' if x <= 2 else 'Neutral')
)

# Monthly analysis with comments
print(f"\n📊 Calculating monthly sentiment breakdown with comments...\n")

monthly_summary = []

for month in sorted(feedback_df['month'].unique()):
    month_data = feedback_df[feedback_df['month'] == month]
    
    total_feedback = len(month_data)
    positive = len(month_data[month_data['sentiment'] == 'Positive'])
    negative = len(month_data[month_data['sentiment'] == 'Negative'])
    neutral = len(month_data[month_data['sentiment'] == 'Neutral'])
    avg_rating = month_data['rating'].mean()
    
    positive_pct = round(positive / total_feedback * 100, 1) if total_feedback > 0 else 0
    negative_pct = round(negative / total_feedback * 100, 1) if total_feedback > 0 else 0
    neutral_pct = round(neutral / total_feedback * 100, 1) if total_feedback > 0 else 0
    
    # Get positive comments
    pos_comments = month_data[month_data['sentiment'] == 'Positive']['comment'].dropna()
    positive_comments_list = '\n'.join(pos_comments.head(5).values) if len(pos_comments) > 0 else "No comments"
    
    # Get negative comments
    neg_comments = month_data[month_data['sentiment'] == 'Negative']['comment'].dropna()
    negative_comments_list = '\n'.join(neg_comments.head(5).values) if len(neg_comments) > 0 else "No comments"
    
    monthly_summary.append({
        'Month': month,
        'Total Feedback': total_feedback,
        'Avg Rating': round(avg_rating, 2),
        'Positive': positive,
        'Positive %': positive_pct,
        'Neutral': neutral,
        'Neutral %': neutral_pct,
        'Negative': negative,
        'Negative %': negative_pct,
        'Positive Comments (Top 5)': positive_comments_list,
        'Negative Comments (Top 5)': negative_comments_list
    })

result = pd.DataFrame(monthly_summary)

# Create Excel with formatting
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=10)

# Sentiment colors
positive_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
negative_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red
neutral_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")   # Yellow

wb = Workbook()
ws = wb.active
ws.title = "Monthly Feedback Summary"

# Write header
for c_idx, col_name in enumerate(result.columns, 1):
    cell = ws.cell(row=1, column=c_idx, value=col_name)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Write data
for r_idx, (_, row) in enumerate(result.iterrows(), 2):
    for c_idx, (col_name, value) in enumerate(row.items(), 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # Apply sentiment colors
        if col_name == 'Positive' or col_name == 'Positive %':
            cell.fill = positive_fill
        elif col_name == 'Negative' or col_name == 'Negative %':
            cell.fill = negative_fill
        elif col_name == 'Neutral' or col_name == 'Neutral %':
            cell.fill = neutral_fill

# Auto-fit columns
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 12
ws.column_dimensions['H'].width = 12
ws.column_dimensions['I'].width = 12
ws.column_dimensions['J'].width = 50  # Positive comments
ws.column_dimensions['K'].width = 50  # Negative comments

ws.freeze_panes = 'A2'
wb.save('data_exports/Monthly_Feedback_Summary_Detailed.xlsx')
print(f"✅ File saved: data_exports/Monthly_Feedback_Summary_Detailed.xlsx\n")

# Print summary table (without comments for readability)
summary_table = result[['Month', 'Total Feedback', 'Avg Rating', 'Positive', 'Positive %', 'Neutral', 'Neutral %', 'Negative', 'Negative %']]
print("="*140)
print("📊 MONTHLY FEEDBACK SENTIMENT ANALYSIS")
print("="*140)
print(summary_table.to_string(index=False))
print("="*140)

# Print detailed comments for each month
print("\n📝 DETAILED FEEDBACK COMMENTS BY MONTH:\n")
for _, row in result.iterrows():
    print(f"\n{'='*140}")
    print(f"📅 {row['Month']} - Total Feedback: {row['Total Feedback']}")
    print(f"   Avg Rating: {row['Avg Rating']}/5.0 | Positive: {row['Positive']} ({row['Positive %']}%) | Negative: {row['Negative']} ({row['Negative %']}%)")
    print(f"{'='*140}")
    
    print(f"\n✅ POSITIVE FEEDBACK ({row['Positive']} comments):")
    print("-" * 140)
    print(row['Positive Comments (Top 5)'])
    
    print(f"\n❌ NEGATIVE FEEDBACK ({row['Negative']} comments):")
    print("-" * 140)
    print(row['Negative Comments (Top 5)'])

# Overall statistics
print(f"\n\n{'='*140}")
print("🎯 OVERALL STATISTICS:")
print(f"{'='*140}")
print(f"   Total Feedback: {result['Total Feedback'].sum():,}")
print(f"   Overall Avg Rating: {feedback_df['rating'].mean():.2f}/5.0")
print(f"   Total Positive Feedback: {result['Positive'].sum():,} ({(result['Positive'].sum() / result['Total Feedback'].sum() * 100):.1f}%)")
print(f"   Total Neutral Feedback: {result['Neutral'].sum():,} ({(result['Neutral'].sum() / result['Total Feedback'].sum() * 100):.1f}%)")
print(f"   Total Negative Feedback: {result['Negative'].sum():,} ({(result['Negative'].sum() / result['Total Feedback'].sum() * 100):.1f}%)")

best_month = result.loc[result['Avg Rating'].idxmax()]
worst_month = result.loc[result['Avg Rating'].idxmin()]
print(f"\n   Best Month: {best_month['Month']} (Avg Rating: {best_month['Avg Rating']}, Positive: {best_month['Positive %']}%)")
print(f"   Worst Month: {worst_month['Month']} (Avg Rating: {worst_month['Avg Rating']}, Negative: {worst_month['Negative %']}%)")
print(f"{'='*140}")
