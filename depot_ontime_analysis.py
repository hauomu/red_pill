import sqlite3
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

print("Generating On-Time Performance Analysis from Database...\n")

# Connect to database for accurate datetime data
conn = sqlite3.connect('delivery.db')
query = """
SELECT 
    deliveries.delivery_id, 
    branch, 
    delivery_datetime, 
    promised_delivery_datetime,
    rating,
    booking_datetime
FROM deliveries 
LEFT JOIN feedback ON deliveries.delivery_id = feedback.delivery_id
"""

df = pd.read_sql_query(query, conn)
conn.close()

print(f"✅ Loaded: {len(df):,} rows from database")

# Parse dates properly (database stores in ISO format)
df['delivery_datetime'] = pd.to_datetime(df['delivery_datetime'], errors='coerce')
df['promised_delivery_datetime'] = pd.to_datetime(df['promised_delivery_datetime'], errors='coerce')
df['booking_datetime'] = pd.to_datetime(df['booking_datetime'], format='%d/%m/%Y %H:%M', errors='coerce')

# Calculate on-time
df['on_time'] = df['delivery_datetime'] <= df['promised_delivery_datetime']

# Extract month from booking
df['month'] = df['booking_datetime'].dt.to_period('M').astype(str)

# Standardize branch names
df['branch'] = df['branch'].str.lower().str.strip()

print(f"✅ Valid datetime records: {df['delivery_datetime'].notna().sum():,}")

print("\n" + "="*140)
print("MONTHLY ON-TIME PERFORMANCE BY DEPOT")
print("="*140)

monthly_ontime = []

for month in sorted(df['month'].dropna().unique()):
    month_data = df[df['month'] == month]
    
    print(f"\n📅 {month}:")
    for depot in sorted(month_data['branch'].unique()):
        if pd.isna(depot):
            continue
        depot_data = month_data[month_data['branch'] == depot]
        
        # Count on-time deliveries (exclude where datetime is null)
        valid_deliveries = depot_data[depot_data['delivery_datetime'].notna() & 
                                      depot_data['promised_delivery_datetime'].notna()]
        
        total = len(valid_deliveries)
        if total == 0:
            continue
            
        on_time_count = valid_deliveries['on_time'].sum()
        on_time_pct = (on_time_count / total * 100)
        
        print(f"  {depot.upper():10} - On-Time: {on_time_count:5} / {total:5} = {on_time_pct:5.1f}%")
        
        monthly_ontime.append({
            'Month': month,
            'Depot': depot.upper(),
            'Total Deliveries': len(depot_data),
            'Valid Date Records': total,
            'On-Time Count': on_time_count,
            'On-Time %': round(on_time_pct, 1)
        })

ontime_result = pd.DataFrame(monthly_ontime)

print("\n" + "="*140)
print("OVERALL DEPOT ON-TIME PERFORMANCE")
print("="*140)

overall_ontime = []
for depot in sorted(df['branch'].dropna().unique()):
    depot_data = df[df['branch'] == depot]
    valid_deliveries = depot_data[depot_data['delivery_datetime'].notna() & 
                                  depot_data['promised_delivery_datetime'].notna()]
    
    total = len(valid_deliveries)
    if total == 0:
        continue
    
    on_time_count = valid_deliveries['on_time'].sum()
    on_time_pct = (on_time_count / total * 100)
    late_count = total - on_time_count
    
    overall_ontime.append({
        'Depot': depot.upper(),
        'Total Deliveries': len(depot_data),
        'Valid Records': total,
        'On-Time': on_time_count,
        'Late': late_count,
        'On-Time %': round(on_time_pct, 1)
    })

overall_result = pd.DataFrame(overall_ontime)
print(overall_result.to_string(index=False))

# Create Excel workbook
wb = Workbook()

# Sheet 1: Monthly On-Time by Depot
ws1 = wb.active
ws1.title = "Monthly On-Time Performance"

thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=10)
good_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green >=90%
warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow 80-90%
bad_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red <80%

for c_idx, col_name in enumerate(ontime_result.columns, 1):
    cell = ws1.cell(row=1, column=c_idx, value=col_name)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

for r_idx, (_, row) in enumerate(ontime_result.iterrows(), 2):
    for c_idx, (col_name, value) in enumerate(row.items(), 1):
        cell = ws1.cell(row=r_idx, column=c_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Color code On-Time %
        if col_name == 'On-Time %':
            if value >= 90:
                cell.fill = good_fill
            elif value >= 80:
                cell.fill = warning_fill
            else:
                cell.fill = bad_fill

for col in ws1.columns:
    max_len = max(len(str(cell.value)) for cell in col if cell.value)
    ws1.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 20)

ws1.freeze_panes = 'A2'

# Sheet 2: Overall On-Time by Depot
ws2 = wb.create_sheet("Overall On-Time Performance")

for c_idx, col_name in enumerate(overall_result.columns, 1):
    cell = ws2.cell(row=1, column=c_idx, value=col_name)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

for r_idx, (_, row) in enumerate(overall_result.iterrows(), 2):
    for c_idx, (col_name, value) in enumerate(row.items(), 1):
        cell = ws2.cell(row=r_idx, column=c_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Color code On-Time %
        if col_name == 'On-Time %':
            if value >= 90:
                cell.fill = good_fill
            elif value >= 80:
                cell.fill = warning_fill
            else:
                cell.fill = bad_fill

for col in ws2.columns:
    max_len = max(len(str(cell.value)) for cell in col if cell.value)
    ws2.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 20)

ws2.freeze_panes = 'A2'

wb.save('data_exports/Depot_OnTime_Performance.xlsx')
print(f"\n✅ File saved: data_exports/Depot_OnTime_Performance.xlsx")
print("\n📊 Sheets created:")
print("   1. Monthly On-Time Performance - On-time % by depot per month")
print("   2. Overall On-Time Performance - Best/worst on-time performers")
