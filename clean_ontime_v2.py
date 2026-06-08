import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import sqlite3

print("=== CLEANING ON-TIME PERFORMANCE DATA (v2) ===\n")

conn = sqlite3.connect('delivery.db')

# Query using WITH to normalize branch names first
query_overall = """
WITH normalized_deliveries AS (
    SELECT 
        CASE 
            WHEN UPPER(branch) IN ('CENTRAL', 'CNETRAL') THEN 'CENTRAL'
            WHEN UPPER(branch) IN ('NORTH', 'NOTH') THEN 'NORTH'
            WHEN UPPER(branch) = 'EAST' THEN 'EAST'
            WHEN UPPER(branch) = 'WEST' THEN 'WEST'
            ELSE UPPER(branch)
        END AS Depot,
        delivery_datetime,
        promised_delivery_datetime
    FROM deliveries
)
SELECT 
    Depot,
    COUNT(*) as Total_Deliveries,
    SUM(CASE WHEN datetime(delivery_datetime) <= datetime(promised_delivery_datetime) THEN 1 ELSE 0 END) as On_Time,
    SUM(CASE WHEN datetime(delivery_datetime) > datetime(promised_delivery_datetime) THEN 1 ELSE 0 END) as Late
FROM normalized_deliveries
GROUP BY Depot
ORDER BY Total_Deliveries DESC
"""

df_overall = pd.read_sql_query(query_overall, conn)
df_overall['On-Time %'] = (df_overall['On_Time'] / df_overall['Total_Deliveries'] * 100).round(1)
df_overall = df_overall.rename(columns={'Total_Deliveries': 'Total Deliveries', 'On_Time': 'On-Time', 'Late': 'Late'})
df_overall = df_overall[['Depot', 'Total Deliveries', 'On-Time', 'Late', 'On-Time %']]
df_overall = df_overall.sort_values('On-Time %', ascending=False).reset_index(drop=True)

print("=== OVERALL ON-TIME PERFORMANCE (CLEANED) ===\n")
print(df_overall.to_string(index=False))

# Query: Monthly on-time performance
query_monthly = """
WITH normalized_deliveries AS (
    SELECT 
        CASE 
            WHEN UPPER(branch) IN ('CENTRAL', 'CNETRAL') THEN 'CENTRAL'
            WHEN UPPER(branch) IN ('NORTH', 'NOTH') THEN 'NORTH'
            WHEN UPPER(branch) = 'EAST' THEN 'EAST'
            WHEN UPPER(branch) = 'WEST' THEN 'WEST'
            ELSE UPPER(branch)
        END AS Depot,
        strftime('%Y-%m', delivery_datetime) AS Month,
        delivery_datetime,
        promised_delivery_datetime
    FROM deliveries
)
SELECT 
    Depot,
    Month,
    COUNT(*) as Total_Deliveries,
    SUM(CASE WHEN datetime(delivery_datetime) <= datetime(promised_delivery_datetime) THEN 1 ELSE 0 END) as On_Time
FROM normalized_deliveries
GROUP BY Depot, Month
ORDER BY Month, Depot
"""

df_monthly_raw = pd.read_sql_query(query_monthly, conn)
df_monthly_raw['On-Time %'] = (df_monthly_raw['On_Time'] / df_monthly_raw['Total_Deliveries'] * 100).round(1)

# Pivot
df_monthly = df_monthly_raw.pivot_table(
    index='Depot',
    columns='Month',
    values='On-Time %',
    aggfunc='first'
)
df_monthly = df_monthly.reset_index()
df_monthly = df_monthly.sort_values('Depot')

print("\n\n=== MONTHLY ON-TIME PERFORMANCE (CLEANED) ===\n")
print(df_monthly.to_string(index=False))

conn.close()

# Create workbook
wb = Workbook()
wb.remove(wb.active)

ws_overall = wb.create_sheet('Overall On-Time Performance')
ws_monthly = wb.create_sheet('Monthly On-Time Performance')

# Formatting
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

def format_sheet(ws, df):
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            if r_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
            else:
                if 'On-Time %' in df.columns and df.columns[c_idx-1] == 'On-Time %':
                    pct = value
                    if pct and pct >= 90:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif pct and pct >= 80:
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

format_sheet(ws_overall, df_overall)
ws_overall.column_dimensions['A'].width = 15
ws_overall.column_dimensions['B'].width = 18
ws_overall.column_dimensions['C'].width = 12
ws_overall.column_dimensions['D'].width = 12
ws_overall.column_dimensions['E'].width = 15

format_sheet(ws_monthly, df_monthly)
ws_monthly.column_dimensions['A'].width = 15
for col in df_monthly.columns[1:]:
    ws_monthly.column_dimensions[ws_monthly.cell(1, list(df_monthly.columns).index(col)+1).column_letter].width = 12

wb.save('data_exports/Depot_OnTime_Performance.xlsx')

print("\n\n✅ CLEANED AND SAVED: data_exports/Depot_OnTime_Performance.xlsx")
print("   ✓ Overall On-Time Performance: 4 main depots (CENTRAL, EAST, NORTH, WEST)")
print("   ✓ Monthly On-Time Performance: 4 main depots by month")
print("\nColor coding: 🟢 ≥90% | 🟡 80-90% | 🔴 <80%")
