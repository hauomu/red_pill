import pandas as pd
import sqlite3
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

print("📖 Reading existing Monthly_Feedback_Enhanced.xlsx...")
existing_df = pd.read_excel('data_exports/Monthly_Feedback_Enhanced.xlsx', sheet_name='Depot Monthly Breakdown')
print(f"Loaded: {len(existing_df)} rows")
print(f"Columns: {list(existing_df.columns)}\n")

# Load on-time data from database
print("📊 Loading on-time percentage data from database...")
conn = sqlite3.connect('delivery.db')

query = """
SELECT 
    strftime('%Y-%m', d.booking_datetime) as Month,
    CASE 
        WHEN UPPER(TRIM(d.branch)) IN ('CENTRAL', 'CNETRAL') THEN 'CENTRAL'
        WHEN UPPER(TRIM(d.branch)) = 'EAST' THEN 'EAST'
        WHEN UPPER(TRIM(d.branch)) IN ('NORTH', 'NOTH') THEN 'NORTH'
        WHEN UPPER(TRIM(d.branch)) = 'WEST' THEN 'WEST'
        ELSE UPPER(TRIM(d.branch))
    END as Depot,
    ROUND(AVG(CASE WHEN d.delivery_datetime <= d.promised_delivery_datetime THEN 1 ELSE 0 END) * 100, 2) as OnTime_Percentage
FROM deliveries d
GROUP BY Month, Depot
ORDER BY Month, Depot
"""

ontime_df = pd.read_sql_query(query, conn)
conn.close()
print(f"Loaded: {len(ontime_df)} rows\n")

# Merge
print("🔗 Merging data...")
result_df = existing_df.merge(ontime_df, on=['Month', 'Depot'], how='left')

print(f"Result: {len(result_df)} rows")
print(f"New columns: {list(result_df.columns)}\n")

# Show preview
print("Preview (first 5 rows):")
print(result_df[['Month', 'Depot', 'Total Deliveries', 'Avg Rating', 'OnTime_Percentage']].head())

# Update the Excel file
print("\n📝 Updating Monthly_Feedback_Enhanced.xlsx...")

wb = load_workbook('data_exports/Monthly_Feedback_Enhanced.xlsx')
ws = wb['Depot Monthly Breakdown']

# Remove old content (keep headers)
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        cell.value = None

# Add new header if not exists
headers = list(result_df.columns)
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.value = header

# Apply header formatting
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                     top=Side(style='thin'), bottom=Side(style='thin'))
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)

for col_idx in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Add data rows
for row_idx, row_data in enumerate(result_df.values, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Format numbers
        if isinstance(value, float):
            cell.number_format = '0.00'

# Auto-fit columns
for col_idx in range(1, len(headers) + 1):
    max_length = 0
    col_letter = get_column_letter(col_idx)
    for row in ws[col_letter]:
        try:
            if row.value and len(str(row.value)) > max_length:
                max_length = len(str(row.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 45)
    ws.column_dimensions[col_letter].width = adjusted_width

# Freeze header
ws.freeze_panes = 'A2'

# Save
wb.save('data_exports/Monthly_Feedback_Enhanced.xlsx')
print("✅ Saved: data_exports/Monthly_Feedback_Enhanced.xlsx\n")

print("=" * 100)
print("✅ UPDATE COMPLETE")
print("=" * 100)
print(f"Total rows: {len(result_df)}")
print(f"New column added: 'OnTime_Percentage'")
print("\nPreview of updated data:")
print(result_df[['Month', 'Depot', 'Total Deliveries', 'Avg Rating', 'OnTime_Percentage', 'Positive %', 'Negative %']].to_string(index=False))
