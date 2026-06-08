import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

print("Generating monthly feedback scores with sentiment analysis...\n")

df = pd.read_csv('data_exports/deliveries_feedback_merged.csv', encoding='latin-1')
print(f"✅ Loaded: {len(df):,} rows")

# Parse dates
df['booking_dt'] = pd.to_datetime(df['booking_datetime'], format='%d/%m/%Y %H:%M', errors='coerce')
df['month'] = df['booking_dt'].dt.to_period('M').astype(str)

# Filter feedback records only
feedback_df = df[df['rating'].notna()].copy()
print(f"✅ Feedback records: {len(feedback_df):,}")

# Categorize sentiment
feedback_df['sentiment'] = feedback_df['rating'].apply(
    lambda x: 'Positive' if x >= 4 else ('Negative' if x <= 2 else 'Neutral')
)

# Monthly analysis
print(f"\n📊 Calculating monthly sentiment breakdown...\n")

monthly_summary = []

for month in sorted(feedback_df['month'].unique()):
    month_data = feedback_df[feedback_df['month'] == month]
    
    total_feedback = len(month_data)
    positive = len(month_data[month_data['sentiment'] == 'Positive'])
    negative = len(month_data[month_data['sentiment'] == 'Negative'])
    neutral = len(month_data[month_data['sentiment'] == 'Neutral'])
    avg_rating = month_data['rating'].mean()
    
    positive_pct = (positive / total_feedback * 100).round(1) if total_feedback > 0 else 0
    negative_pct = (negative / total_feedback * 100).round(1) if total_feedback > 0 else 0
    neutral_pct = (neutral / total_feedback * 100).round(1) if total_feedback > 0 else 0
    
    # Get sample positive comment
    pos_comments = month_data[month_data['sentiment'] == 'Positive']['comment'].dropna()
    sample_positive = pos_comments.iloc[0][:80] if len(pos_comments) > 0 else "N/A"
    
    # Get sample negative comment
    neg_comments = month_data[month_data['sentiment'] == 'Negative']['comment'].dropna()
    sample_negative = neg_comments.iloc[0][:80] if len(neg_comments) > 0 else "N/A"
    
    monthly_summary.append({
        'Month': month,
        'Total Feedback': total_feedback,
        'Avg Rating': round(avg_rating, 2),
        'Positive': positive,
        'Positive %': positive_pct,
        'Neutral': neutral,
        'Neutral %': neutral_pct,
        'Negative': negative,
        'Negative %': negative_pct,
        'Sample Positive': sample_positive,
        'Sample Negative': sample_negative
    })

result = pd.DataFrame(monthly_summary)

# Create Excel with formatting
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=10)

# Sentiment colors
positive_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
negative_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red
neutral_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")   # Yellow

wb = Workbook()
ws = wb.active
ws.title = "Monthly Sentiment"

# Write header
for c_idx, col_name in enumerate(result.columns, 1):
    cell = ws.cell(row=1, column=c_idx, value=col_name)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Write data
for r_idx, (_, row) in enumerate(result.iterrows(), 2):
    for c_idx, (col_name, value) in enumerate(row.items(), 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Apply sentiment colors
        if col_name == 'Positive' or col_name == 'Positive %':
            cell.fill = positive_fill
        elif col_name == 'Negative' or col_name == 'Negative %':
            cell.fill = negative_fill
        elif col_name == 'Neutral' or col_name == 'Neutral %':
            cell.fill = neutral_fill

# Auto-fit columns
for col in ws.columns:
    max_len = max(len(str(cell.value)) for cell in col if cell.value)
    col_letter = get_column_letter(col[0].column)
    ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

ws.freeze_panes = 'A2'
wb.save('data_exports/Monthly_Feedback_Sentiment.xlsx')
print(f"✅ File saved: data_exports/Monthly_Feedback_Sentiment.xlsx\n")

# Print summary
print("="*140)
print("📊 MONTHLY FEEDBACK SENTIMENT ANALYSIS")
print("="*140)
print(result.to_string(index=False))
print("="*140)

# Overall statistics
print("\n🎯 OVERALL STATISTICS:")
print(f"   Total Feedback: {result['Total Feedback'].sum():,}")
print(f"   Overall Avg Rating: {feedback_df['rating'].mean():.2f}/5.0")
print(f"   Total Positive Feedback: {result['Positive'].sum():,} ({(result['Positive'].sum() / result['Total Feedback'].sum() * 100):.1f}%)")
print(f"   Total Neutral Feedback: {result['Neutral'].sum():,} ({(result['Neutral'].sum() / result['Total Feedback'].sum() * 100):.1f}%)")
print(f"   Total Negative Feedback: {result['Negative'].sum():,} ({(result['Negative'].sum() / result['Total Feedback'].sum() * 100):.1f}%)")

# Best and worst months
best_month = result.loc[result['Avg Rating'].idxmax()]
worst_month = result.loc[result['Avg Rating'].idxmin()]
print(f"\n   Best Month: {best_month['Month']} (Avg Rating: {best_month['Avg Rating']}, Positive: {best_month['Positive %']}%)")
print(f"   Worst Month: {worst_month['Month']} (Avg Rating: {worst_month['Avg Rating']}, Negative: {worst_month['Negative %']}%)")
print("="*140)
