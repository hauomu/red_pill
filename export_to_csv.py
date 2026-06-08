import sqlite3
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

conn = sqlite3.connect('delivery.db')

# Create exports folder
os.makedirs('data_exports', exist_ok=True)

print("Exporting database tables with formatted borders...\n")

# Define border style
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

def export_to_excel(df, filename, title=""):
    """Export dataframe to Excel with borders and formatting"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    
    # Add title if provided
    if title:
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=14)
        start_row = 3
    else:
        start_row = 1
    
    # Write dataframe to Excel
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Format header row
            if r_idx == start_row:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze header row
    ws.freeze_panes = f'A{start_row + 1}'
    
    wb.save(filename)
    print(f"   ✅ Saved: {filename}")

# 1. Deliveries Sample
print("📦 Exporting deliveries table...")
df_deliveries = pd.read_sql_query("SELECT * FROM deliveries LIMIT 1000", conn)
export_to_excel(df_deliveries, 'data_exports/deliveries_sample.xlsx', '📦 Deliveries Sample (1000 rows)')

# 2. Deliveries by Branch Stats
print("📊 Exporting deliveries statistics...")
df_deliv_stats = pd.read_sql_query("""
SELECT 
    branch, 
    COUNT(*) as total_deliveries,
    ROUND(AVG(CASE WHEN delivery_datetime <= promised_delivery_datetime THEN 1 ELSE 0 END) * 100, 1) as on_time_pct,
    ROUND(AVG(distance_km), 2) as avg_distance_km,
    ROUND(AVG(CAST(num_stops_on_route as FLOAT)), 1) as avg_stops,
    ROUND(AVG(driver_experience_months), 1) as avg_driver_exp_months
FROM deliveries
WHERE branch IN ('Central', 'East', 'West', 'North')
GROUP BY branch
ORDER BY total_deliveries DESC
""", conn)
export_to_excel(df_deliv_stats, 'data_exports/deliveries_by_branch.xlsx', '📊 Deliveries by Branch')

# 3. Feedback Sample
print("\n⭐ Exporting feedback table...")
df_feedback = pd.read_sql_query("SELECT * FROM feedback LIMIT 1000", conn)
export_to_excel(df_feedback, 'data_exports/feedback_sample.xlsx', '⭐ Feedback Sample (1000 rows)')

# 4. Feedback Rating Distribution
print("📊 Exporting feedback statistics...")
df_feedback_stats = pd.read_sql_query("""
SELECT 
    rating,
    COUNT(*) as count,
    ROUND(COUNT(*) * 100.0 / (SELECT COUNT(*) FROM feedback WHERE rating IS NOT NULL), 1) as percentage
FROM feedback
WHERE rating IS NOT NULL
GROUP BY rating
ORDER BY rating
""", conn)
export_to_excel(df_feedback_stats, 'data_exports/feedback_rating_distribution.xlsx', '⭐ Feedback Rating Distribution')

# 5. Delivery vs Feedback Correlation
print("🔗 Exporting delivery-feedback correlation...")
df_correlation = pd.read_sql_query("""
SELECT 
    CASE WHEN d.delivery_datetime <= d.promised_delivery_datetime THEN 'On-time' ELSE 'Late' END as status,
    AVG(f.rating) as avg_rating,
    COUNT(f.rating) as feedback_count,
    MIN(f.rating) as min_rating,
    MAX(f.rating) as max_rating
FROM deliveries d
LEFT JOIN feedback f ON d.delivery_id = f.delivery_id
WHERE f.rating IS NOT NULL
GROUP BY status
""", conn)
export_to_excel(df_correlation, 'data_exports/delivery_status_vs_rating.xlsx', '🔗 Delivery Status vs Rating')

# 6. Vehicle Type Analysis
print("🚚 Exporting vehicle analysis...")
df_vehicles = pd.read_sql_query("""
SELECT 
    vehicle_type,
    COUNT(*) as total,
    ROUND(COUNT(*) * 100.0 / (SELECT COUNT(*) FROM deliveries WHERE vehicle_type IN ('van', 'bike', 'truck')), 1) as percentage,
    ROUND(AVG(CASE WHEN delivery_datetime <= promised_delivery_datetime THEN 1 ELSE 0 END) * 100, 1) as on_time_pct,
    ROUND(AVG(distance_km), 2) as avg_distance
FROM deliveries
WHERE vehicle_type IN ('van', 'bike', 'truck')
GROUP BY vehicle_type
ORDER BY total DESC
""", conn)
export_to_excel(df_vehicles, 'data_exports/vehicle_analysis.xlsx', '🚚 Vehicle Type Analysis')

# 7. Priority Level Analysis
print("📍 Exporting priority analysis...")
df_priority = pd.read_sql_query("""
SELECT 
    delivery_priority,
    COUNT(*) as total,
    ROUND(AVG(CASE WHEN delivery_datetime <= promised_delivery_datetime THEN 1 ELSE 0 END) * 100, 1) as on_time_pct,
    ROUND(AVG(distance_km), 2) as avg_distance,
    ROUND(AVG(CAST(parcel_weight_kg as FLOAT)), 2) as avg_weight
FROM deliveries
GROUP BY delivery_priority
ORDER BY total DESC
""", conn)
export_to_excel(df_priority, 'data_exports/priority_analysis.xlsx', '📍 Priority Level Analysis')

# Summary
print("\n" + "="*60)
print("EXPORT SUMMARY")
print("="*60)

summary = f"""
📁 ALL FILES EXPORTED TO: data_exports/

📊 EXCEL FILES (with borders & formatting):
   ✅ deliveries_sample.xlsx (1000 rows)
   ✅ deliveries_by_branch.xlsx (statistics)
   ✅ feedback_sample.xlsx (1000 rows)
   ✅ feedback_rating_distribution.xlsx (statistics)
   ✅ delivery_status_vs_rating.xlsx (correlation)
   ✅ vehicle_analysis.xlsx (by vehicle type)
   ✅ priority_analysis.xlsx (by priority level)

✨ FEATURES:
   • All cells have borders (row & column lines)
   • Header rows formatted with blue background
   • Auto-adjusted column widths
   • Center-aligned text
   • Frozen header rows for easy scrolling
   • Open directly in Excel

📂 LOCATIONS:
   • Windows: data_exports/
   • Open with: Microsoft Excel or Google Sheets
"""

print(summary)

conn.close()
print("\n✅ All exports complete! Open .xlsx files to see formatted tables.\n")
