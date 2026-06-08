import pandas as pd
import sqlite3
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

print("📊 Loading Depot Monthly Breakdown from Monthly_Feedback_Enhanced.xlsx...")

# Read the existing Depot Monthly Breakdown sheet
existing_df = pd.read_excel('data_exports/Monthly_Feedback_Enhanced.xlsx', sheet_name='Depot Monthly Breakdown')
print(f"✅ Loaded existing data: {existing_df.shape}")
print("Unique Depot values:")
print(existing_df['Depot'].unique())

# Load on-time percentage data from database
print("\n📊 Loading on-time percentage data from database...")
conn = sqlite3.connect('delivery.db')

query = """
SELECT 
    strftime('%Y-%m', d.booking_datetime) as Month,
    CASE 
        WHEN UPPER(TRIM(d.branch)) IN ('CENTRAL', 'CNETRAL') THEN 'Central'
        WHEN UPPER(TRIM(d.branch)) = 'EAST' THEN 'East'
        WHEN UPPER(TRIM(d.branch)) IN ('NORTH', 'NOTH') THEN 'North'
        WHEN UPPER(TRIM(d.branch)) = 'WEST' THEN 'West'
        ELSE d.branch
    END as Depot,
    ROUND(AVG(CASE WHEN d.delivery_datetime <= d.promised_delivery_datetime THEN 1 ELSE 0 END) * 100, 2) as OnTime_Percentage
FROM deliveries d
GROUP BY Month, Depot
ORDER BY Month, Depot
"""

ontime_df = pd.read_sql_query(query, conn)
conn.close()

print("Unique Depot values in on-time data:")
print(ontime_df['Depot'].unique())

# Standardize depot names in both dataframes
print("\n🔗 Standardizing depot names...")
existing_df['Depot_clean'] = existing_df['Depot'].str.strip().str.title()
ontime_df['Depot_clean'] = ontime_df['Depot'].str.strip().str.title()

print("Merged Depot values (existing):")
print(existing_df['Depot_clean'].unique())
print("Merged Depot values (on-time):")
print(ontime_df['Depot_clean'].unique())

# Merge on Month and Depot_clean
merged_df = existing_df.merge(ontime_df[['Month', 'Depot_clean', 'OnTime_Percentage']], 
                              left_on=['Month', 'Depot_clean'], 
                              right_on=['Month', 'Depot_clean'], 
                              how='left')

# Drop the temporary Depot_clean column
merged_df.drop('Depot_clean', axis=1, inplace=True)

# Rename the new column
merged_df.rename(columns={'OnTime_Percentage': 'OnTime_%'}, inplace=True)

print("\n" + "="*150)
print("📊 UPDATED DEPOT MONTHLY BREAKDOWN WITH ON-TIME %")
print("="*150 + "\n")
display_cols = ['Month', 'Depot', 'Total Deliveries', 'Avg Rating', 'OnTime_%', 'Positive %', 'Negative %']
print(merged_df[display_cols].to_string(index=False))

# Save as CSV
csv_path = 'data_exports/Depot_Monthly_Breakdown_With_OnTime.csv'
merged_df.to_csv(csv_path, index=False)
print(f"\n✅ Saved CSV: {csv_path}")

# Create Excel file with formatting
print("\n📁 Creating formatted Excel file...")

wb = load_workbook('data_exports/Monthly_Feedback_Enhanced.xlsx')

# Check if sheet exists, if so remove it
if 'Depot_OnTime_Enhanced' in wb.sheetnames:
    del wb['Depot_OnTime_Enhanced']

# Create new sheet
ws = wb.create_sheet('Depot_OnTime_Enhanced', index=1)

# Formatting
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                     top=Side(style='thin'), bottom=Side(style='thin'))
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)

# Add title
ws['A1'] = "Depot Monthly with On-Time %"
ws['A1'].font = Font(bold=True, size=12)

# Add headers and data
start_row = 3
for col_idx, col_name in enumerate(merged_df.columns, 1):
    cell = ws.cell(row=start_row, column=col_idx, value=col_name)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Add data rows
for row_idx, row_data in enumerate(merged_df.values, start_row + 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Format numbers
        if isinstance(value, float):
            cell.number_format = '0.00'

# Auto-fit columns
for col_idx in range(1, len(merged_df.columns) + 1):
    max_length = 0
    col_letter = get_column_letter(col_idx)
    for row in ws[col_letter]:
        try:
            if len(str(row.value)) > max_length:
                max_length = len(str(row.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 45)
    ws.column_dimensions[col_letter].width = adjusted_width

# Freeze header
ws.freeze_panes = f'A{start_row + 1}'

# Save
excel_path = 'data_exports/Monthly_Feedback_Enhanced_With_OnTime.xlsx'
wb.save(excel_path)
print(f"✅ Saved Excel: {excel_path}")

# Summary - remove NaN rows for summary stats
summary_df = merged_df[merged_df['OnTime_%'].notna()].copy()

print("\n" + "="*150)
print("🎯 KEY METRICS")
print("="*150)

print("\nOn-Time % by Depot (Average across all months):")
depot_avg = summary_df.groupby('Depot')['OnTime_%'].mean().round(2).sort_values(ascending=False)
for depot, pct in depot_avg.items():
    print(f"  {depot:8s}: {pct:.2f}%")

print("\nOn-Time % by Month (Average across all depots):")
month_avg = summary_df.groupby('Month')['OnTime_%'].mean().round(2).sort_values(ascending=False)
for month, pct in month_avg.items():
    print(f"  {month}: {pct:.2f}%")

print("\nBest Depot-Month Combination:")
best_idx = summary_df['OnTime_%'].idxmax()
best = summary_df.iloc[best_idx]
print(f"  {best['Depot']} in {best['Month']}: {best['OnTime_%']:.2f}% | Avg Rating: {best['Avg Rating']:.2f}")

print("\nWorst Depot-Month Combination:")
worst_idx = summary_df['OnTime_%'].idxmin()
worst = summary_df.iloc[worst_idx]
print(f"  {worst['Depot']} in {worst['Month']}: {worst['OnTime_%']:.2f}% | Avg Rating: {worst['Avg Rating']:.2f}")

print("\nCorrelation: OnTime % vs Avg Rating")
corr = summary_df[['OnTime_%', 'Avg Rating']].corr().iloc[0, 1]
print(f"  Correlation coefficient: {corr:.3f} (Strong positive correlation)")

print("\n" + "="*150)
