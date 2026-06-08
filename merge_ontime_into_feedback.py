import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Read both dataframes
df_ontime = pd.read_excel('data_exports/Depot_OnTime_Performance.xlsx', sheet_name='Overall On-Time Performance')
df_feedback = pd.read_excel('data_exports/Monthly_Feedback_Enhanced.xlsx', sheet_name='Overall Depot Performance')

# Clean depot names (standardize - remove leading/trailing spaces)
df_ontime['Depot'] = df_ontime['Depot'].str.strip()
df_feedback['Depot'] = df_feedback['Depot'].str.strip()

# Merge on Depot (keep only main 4 depots for clarity)
main_depots = ['CENTRAL', 'EAST', 'NORTH', 'WEST']
df_ontime_filtered = df_ontime[df_ontime['Depot'].isin(main_depots)].copy()
df_merged = df_feedback.merge(df_ontime_filtered[['Depot', 'On-Time %']], on='Depot', how='left')

# Reorder columns for better readability
df_merged = df_merged[['Depot', 'Total Deliveries', 'Avg Rating', 'On-Time %', 'Min Rating', 'Max Rating']]

# Sort by on-time % descending (best performers first)
df_merged = df_merged.sort_values('On-Time %', ascending=False, na_position='last').reset_index(drop=True)

print("=== MERGED OVERALL DEPOT PERFORMANCE ===")
print(df_merged.to_string(index=False))

# Load workbook and update sheet
wb = load_workbook('data_exports/Monthly_Feedback_Enhanced.xlsx')
ws = wb['Overall Depot Performance']

# Clear existing data (keep header format)
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        cell.value = None

# Write merged data
for r_idx, row in enumerate(dataframe_to_rows(df_merged, index=False, header=False), start=2):
    for c_idx, value in enumerate(row, start=1):
        ws.cell(row=r_idx, column=c_idx, value=value)

# Apply formatting (colors based on rating performance)
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Format headers
for cell in ws[1]:
    if cell.value:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

# Format data rows with color coding
for row_idx in range(2, ws.max_row + 1):
    depot = ws.cell(row=row_idx, column=1).value
    avg_rating = ws.cell(row=row_idx, column=3).value
    ontime_pct = ws.cell(row=row_idx, column=4).value
    
    for col_idx in range(1, 7):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Color code: Green (high) -> Yellow (medium) -> Red (low)
        if col_idx == 3:  # Avg Rating
            if avg_rating and avg_rating >= 4.3:
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif avg_rating and avg_rating >= 4.1:
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        elif col_idx == 4:  # On-Time %
            if ontime_pct and ontime_pct >= 90:
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif ontime_pct and ontime_pct >= 80:
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# Adjust column widths
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 15

# Save updated workbook
wb.save('data_exports/Monthly_Feedback_Enhanced.xlsx')
print("\n✅ Successfully merged 'Overall On-Time Performance' into 'Overall Depot Performance'")
print("✅ File saved: data_exports/Monthly_Feedback_Enhanced.xlsx")
print("\nColor coding applied:")
print("  🟢 Green:  Avg Rating ≥4.3, On-Time % ≥90%")
print("  🟡 Yellow: Avg Rating 4.1-4.3, On-Time % 80-90%")
print("  🔴 Red:    Avg Rating <4.1, On-Time % <80%")
