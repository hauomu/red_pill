import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

print("=== CLEANING ON-TIME PERFORMANCE DATA ===\n")

# Read the raw on-time data from database
import sqlite3

conn = sqlite3.connect('delivery.db')

# Get on-time stats by depot (direct from database)
query = """
SELECT 
    CASE 
        WHEN branch IN ('central', 'cnetral') THEN 'CENTRAL'
        WHEN branch IN ('north', 'noth') THEN 'NORTH'
        WHEN branch = 'east' THEN 'EAST'
        WHEN branch = 'west' THEN 'WEST'
        ELSE UPPER(branch)
    END AS Depot,
    COUNT(*) as Total_Deliveries,
    SUM(CASE WHEN datetime(delivery_datetime) <= datetime(promised_delivery_datetime) THEN 1 ELSE 0 END) as On_Time,
    SUM(CASE WHEN datetime(delivery_datetime) > datetime(promised_delivery_datetime) THEN 1 ELSE 0 END) as Late
FROM deliveries
GROUP BY Depot
ORDER BY Total_Deliveries DESC
"""

df_clean = pd.read_sql_query(query, conn)
conn.close()

# Calculate on-time percentage
df_clean['On-Time %'] = (df_clean['On_Time'] / df_clean['Total_Deliveries'] * 100).round(1)

# Rename columns for clarity
df_clean = df_clean.rename(columns={
    'Total_Deliveries': 'Total Deliveries',
    'On_Time': 'On-Time',
    'Late': 'Late'
})

# Reorder columns
df_clean = df_clean[['Depot', 'Total Deliveries', 'On-Time', 'Late', 'On-Time %']]

# Sort by on-time % descending
df_clean = df_clean.sort_values('On-Time %', ascending=False).reset_index(drop=True)

print("=== CLEANED OVERALL ON-TIME PERFORMANCE ===\n")
print(df_clean.to_string(index=False))
print("\n✅ Merged typo depots (CNETRAL→CENTRAL, NOTH→NORTH)")

# Load workbook and update both sheets
wb = load_workbook('data_exports/Depot_OnTime_Performance.xlsx')

# Clean up Overall On-Time Performance sheet
ws = wb['Overall On-Time Performance']
ws.delete_rows(2, ws.max_row)  # Delete all data rows

# Write cleaned data
for r_idx, row in enumerate(dataframe_to_rows(df_clean, index=False, header=False), start=2):
    for c_idx, value in enumerate(row, start=1):
        ws.cell(row=r_idx, column=c_idx, value=value)

# Format the sheet
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Format headers
for cell in ws[1]:
    if cell.value:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

# Format data rows with color coding
for row_idx in range(2, ws.max_row + 1):
    ontime_pct = ws.cell(row=row_idx, column=5).value
    
    for col_idx in range(1, 6):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Color code On-Time % column
        if col_idx == 5:
            if ontime_pct and ontime_pct >= 90:
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif ontime_pct and ontime_pct >= 80:
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# Adjust column widths
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 15

# Also clean Monthly On-Time Performance sheet (filter to 4 main depots)
ws_monthly = wb['Monthly On-Time Performance']

# Read the monthly data
df_monthly = pd.read_excel('data_exports/Depot_OnTime_Performance.xlsx', sheet_name='Monthly On-Time Performance')

# Standardize depot names
df_monthly['Depot'] = df_monthly['Depot'].str.strip().str.upper()
df_monthly['Depot'] = df_monthly['Depot'].replace({
    'CNETRAL': 'CENTRAL',
    'NOTH': 'NORTH'
})

# Keep only main 4 depots
main_depots = ['CENTRAL', 'EAST', 'NORTH', 'WEST']
df_monthly = df_monthly[df_monthly['Depot'].isin(main_depots)].copy()

# Sort by Depot name
df_monthly = df_monthly.sort_values('Depot').reset_index(drop=True)

# Clear and rewrite monthly sheet
ws_monthly.delete_rows(2, ws_monthly.max_row)

for r_idx, row in enumerate(dataframe_to_rows(df_monthly, index=False, header=False), start=2):
    for c_idx, value in enumerate(row, start=1):
        ws_monthly.cell(row=r_idx, column=c_idx, value=value)

# Format monthly sheet
for cell in ws_monthly[1]:
    if cell.value:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

for row_idx in range(2, ws_monthly.max_row + 1):
    for col_idx in range(1, ws_monthly.max_column + 1):
        cell = ws_monthly.cell(row=row_idx, column=col_idx)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

ws_monthly.column_dimensions['A'].width = 15
for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
    ws_monthly.column_dimensions[col].width = 12

# Save cleaned workbook
wb.save('data_exports/Depot_OnTime_Performance.xlsx')

print("\n✅ File cleaned and saved: data_exports/Depot_OnTime_Performance.xlsx")
print("   • Overall On-Time Performance: Merged typos, kept 4 main depots")
print("   • Monthly On-Time Performance: Filtered to 4 main depots only")
print("\nColor coding:")
print("  🟢 Green:  On-Time % ≥90%")
print("  🟡 Yellow: On-Time % 80-90%")
print("  🔴 Red:    On-Time % <80%")
