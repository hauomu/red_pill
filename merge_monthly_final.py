import pandas as pd
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

print("📊 Loading data from database...")

# Load from database
conn = sqlite3.connect('delivery.db')

# Query: Monthly feedback metrics
feedback_query = """
SELECT 
    strftime('%Y-%m', f.feedback_datetime) as Month,
    COUNT(DISTINCT f.feedback_id) as Total_Feedback,
    ROUND(AVG(f.rating), 2) as Avg_Rating,
    MIN(f.rating) as Min_Rating,
    MAX(f.rating) as Max_Rating,
    ROUND(AVG(CASE WHEN f.rating <= 2 THEN 1 ELSE 0 END) * 100, 1) as Negative_Feedback_Pct,
    ROUND(AVG(CASE WHEN f.rating >= 4 THEN 1 ELSE 0 END) * 100, 1) as Positive_Feedback_Pct
FROM feedback f
GROUP BY Month
ORDER BY Month
"""

feedback_df = pd.read_sql_query(feedback_query, conn)
print(f"✅ Loaded feedback: {len(feedback_df)} months")

# Query: Depot monthly on-time delivery
depot_query = """
SELECT 
    strftime('%Y-%m', d.booking_datetime) as Month,
    CASE 
        WHEN UPPER(TRIM(d.branch)) IN ('CENTRAL', 'CNETRAL') THEN 'Central'
        WHEN UPPER(TRIM(d.branch)) = 'EAST' THEN 'East'
        WHEN UPPER(TRIM(d.branch)) IN ('NORTH', 'NOTH') THEN 'North'
        WHEN UPPER(TRIM(d.branch)) = 'WEST' THEN 'West'
        ELSE d.branch
    END as Depot,
    COUNT(d.delivery_id) as Total_Deliveries,
    ROUND(AVG(CASE WHEN d.delivery_datetime <= d.promised_delivery_datetime THEN 1 ELSE 0 END) * 100, 1) as OnTime_Pct
FROM deliveries d
GROUP BY Month, Depot
ORDER BY Month, Depot
"""

depot_df = pd.read_sql_query(depot_query, conn)
print(f"✅ Loaded depot data: {len(depot_df)} rows")

conn.close()

# Pivot depot data (depots as columns)
depot_pivot = depot_df.pivot_table(
    index='Month',
    columns='Depot',
    values='OnTime_Pct',
    aggfunc='first'
).reset_index()

print("\n📋 Data Preview:")
print("\nFeedback by month:")
print(feedback_df.to_string(index=False))
print("\nDepot On-Time by month:")
print(depot_pivot.to_string(index=False))

# Merge on Month
print("\n🔗 Merging datasets...")
merged_df = feedback_df.merge(depot_pivot, on='Month', how='outer')
merged_df = merged_df.sort_values('Month').reset_index(drop=True)

print("\n" + "="*140)
print("📊 MERGED: MONTHLY FEEDBACK + DEPOT ON-TIME PROPORTION")
print("="*140 + "\n")
print(merged_df.to_string(index=False))

# Save as CSV
csv_path = 'data_exports/Monthly_Feedback_Depot_OnTime_Merged.csv'
merged_df.to_csv(csv_path, index=False)
print(f"\n✅ Saved CSV: {csv_path}")

# Save as Excel with formatting
print("\n📁 Creating Excel file...")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)

wb = Workbook()
ws = wb.active
ws.title = "Monthly Merged"

# Add title
ws['A1'] = "Monthly Feedback & Depot On-Time Performance"
ws['A1'].font = Font(bold=True, size=12)

# Add data starting from row 3
start_row = 3
for r_idx, row in enumerate(dataframe_to_rows(merged_df, index=False, header=True), start_row):
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        cell.border = thin_border
        if r_idx == start_row:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        else:
            cell.alignment = Alignment(horizontal='center')
            if isinstance(value, float):
                cell.number_format = '0.0'
            elif isinstance(value, int):
                cell.number_format = '0'

# Auto-fit columns
for col_idx, col in enumerate(ws.columns, 1):
    max_length = max(len(str(cell.value)) for cell in col if cell.value)
    ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 40)

ws.freeze_panes = f'A{start_row + 1}'

excel_path = 'data_exports/Monthly_Feedback_Depot_OnTime_Merged.xlsx'
wb.save(excel_path)
print(f"✅ Saved Excel: {excel_path}\n")

# Summary statistics
print("="*140)
print("🎯 KEY INSIGHTS")
print("="*140)

print("\nMonthly Trend:")
print("Month | Avg Rating | Central % | East % | North % | West % | Negative % | Positive %")
for _, row in merged_df.iterrows():
    month = row['Month']
    rating = f"{row['Avg_Rating']:.2f}" if pd.notna(row['Avg_Rating']) else "-"
    central = f"{row.get('Central', '-'):.1f}" if pd.notna(row.get('Central')) else "-"
    east = f"{row.get('East', '-'):.1f}" if pd.notna(row.get('East')) else "-"
    north = f"{row.get('North', '-'):.1f}" if pd.notna(row.get('North')) else "-"
    west = f"{row.get('West', '-'):.1f}" if pd.notna(row.get('West')) else "-"
    neg = f"{row.get('Negative_Feedback_Pct', '-'):.1f}" if pd.notna(row.get('Negative_Feedback_Pct')) else "-"
    pos = f"{row.get('Positive_Feedback_Pct', '-'):.1f}" if pd.notna(row.get('Positive_Feedback_Pct')) else "-"
    print(f"{month} | {rating:>10} | {central:>9} | {east:>6} | {north:>7} | {west:>6} | {neg:>10} | {pos:>10}")

print("\nBest Performing Month (by rating): " + merged_df.loc[merged_df['Avg_Rating'].idxmax(), 'Month'] + 
      f" ({merged_df['Avg_Rating'].max():.2f})")
print("Worst Performing Month (by rating): " + merged_df.loc[merged_df['Avg_Rating'].idxmin(), 'Month'] + 
      f" ({merged_df['Avg_Rating'].min():.2f})")

print("\n" + "="*140)
