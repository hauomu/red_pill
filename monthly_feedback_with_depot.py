import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import re

print("Generating monthly feedback analysis with depot location...\n")

df = pd.read_csv('data_exports/deliveries_feedback_merged.csv', encoding='latin-1')
print(f"✅ Loaded: {len(df):,} rows")

# Parse dates
df['booking_dt'] = pd.to_datetime(df['booking_datetime'], format='%d/%m/%Y %H:%M', errors='coerce')
df['month'] = df['booking_dt'].dt.to_period('M').astype(str)

# Filter feedback records
feedback_df = df[df['rating'].notna()].copy()
print(f"✅ Feedback records: {len(feedback_df):,}\n")

# Categorize sentiment
feedback_df['sentiment'] = feedback_df['rating'].apply(
    lambda x: 'Positive' if x >= 4 else ('Negative' if x <= 2 else 'Neutral')
)

# Complaint categories
complaint_patterns = {
    'Late Delivery': r'(late|slow|hours late|promised window|delay)',
    'Poor Communication': r'(communication|no (update|notification)|not informed)',
    'Driver Behavior': r'(rude|unprofessional|careless|rough)',
    'Package Damage': r'(damage|broken|damaged|roughly handled)',
    'Other': r'.*'
}

def categorize_complaint(comment):
    if pd.isna(comment):
        return 'No Comment'
    comment_lower = str(comment).lower()
    for category, pattern in complaint_patterns.items():
        if re.search(pattern, comment_lower):
            return category
    return 'Other'

negative_feedback = feedback_df[feedback_df['sentiment'] == 'Negative'].copy()
negative_feedback['complaint_category'] = negative_feedback['comment'].apply(categorize_complaint)

print("=" * 140)
print("MONTHLY ANALYSIS BY DEPOT LOCATION (BRANCH)")
print("=" * 140)

monthly_depot_data = []

for month in sorted(feedback_df['month'].unique()):
    month_data = feedback_df[feedback_df['month'] == month]
    
    print(f"\n📅 {month}:")
    for depot in sorted(month_data['branch'].unique()):
        depot_data = month_data[month_data['branch'] == depot]
        if len(depot_data) > 0:
            total = len(depot_data)
            avg_rating = depot_data['rating'].mean()
            std_rating = depot_data['rating'].std()
            neg_count = len(depot_data[depot_data['sentiment'] == 'Negative'])
            neg_pct = (neg_count / total * 100)
            pos_count = len(depot_data[depot_data['sentiment'] == 'Positive'])
            pos_pct = (pos_count / total * 100)
            neutral_count = len(depot_data[depot_data['sentiment'] == 'Neutral'])
            neutral_pct = (neutral_count / total * 100)
            
            print(f"  {depot.upper():10} - Deliveries: {total:5} | Avg Rating: {avg_rating:.2f}±{std_rating:.2f} | Positive: {pos_pct:5.1f}% | Neutral: {neutral_pct:5.1f}% | Negative: {neg_pct:5.1f}%")
            
            monthly_depot_data.append({
                'Month': month,
                'Depot': depot.upper(),
                'Total Deliveries': total,
                'Avg Rating': round(avg_rating, 2),
                'Service Consistency (Std Dev)': round(std_rating, 2),
                'Positive': pos_count,
                'Positive %': round(pos_pct, 1),
                'Neutral': neutral_count,
                'Neutral %': round(neutral_pct, 1),
                'Negative': neg_count,
                'Negative %': round(neg_pct, 1)
            })

depot_result = pd.DataFrame(monthly_depot_data)

# Calculate monthly totals and add percentage
monthly_totals = depot_result.groupby('Month')['Total Deliveries'].transform('sum')
depot_result['Proportion of Monthly Deliveries (%)'] = (depot_result['Total Deliveries'] / monthly_totals * 100).round(1)

# Add grand total row
grand_total_deliveries = depot_result['Total Deliveries'].sum()
depot_result.loc[len(depot_result)] = {
    'Month': 'GRAND TOTAL',
    'Depot': '',
    'Total Deliveries': grand_total_deliveries,
    'Proportion of Monthly Deliveries (%)': 100.0,
    'Avg Rating': '',
    'Service Consistency (Std Dev)': '',
    'Positive': '',
    'Positive %': '',
    'Neutral': '',
    'Neutral %': '',
    'Negative': '',
    'Negative %': ''
}

# Overall depot performance
print("\n" + "=" * 140)
print("OVERALL DEPOT PERFORMANCE")
print("=" * 140)
overall_depot = feedback_df.groupby('branch').agg({
    'delivery_id': 'count',
    'rating': ['mean', 'min', 'max']
}).round(2)
overall_depot.columns = ['Total Deliveries', 'Avg Rating', 'Min Rating', 'Max Rating']
overall_depot = overall_depot.sort_values('Avg Rating', ascending=False)
overall_depot.index = overall_depot.index.str.upper()
print(overall_depot.to_string())

# Create comprehensive Excel with multiple sheets
wb = Workbook()

# Sheet 1: Depot Breakdown by Month
ws1 = wb.active
ws1.title = "Depot Monthly Breakdown"

thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=10)
negative_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
positive_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
neutral_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

for c_idx, col_name in enumerate(depot_result.columns, 1):
    cell = ws1.cell(row=1, column=c_idx, value=col_name)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Reorder columns for better readability (removed 'Monthly Total Deliveries')
depot_result = depot_result[['Month', 'Depot', 'Total Deliveries', 'Proportion of Monthly Deliveries (%)', 
                               'Avg Rating', 'Service Consistency (Std Dev)', 
                               'Positive', 'Positive %', 'Neutral', 'Neutral %', 'Negative', 'Negative %']]

# Rewrite headers with reordered columns
ws1.delete_rows(1)
for c_idx, col_name in enumerate(depot_result.columns, 1):
    cell = ws1.cell(row=1, column=c_idx, value=col_name)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

for r_idx, (_, row) in enumerate(depot_result.iterrows(), 2):
    for c_idx, (col_name, value) in enumerate(row.items(), 1):
        cell = ws1.cell(row=r_idx, column=c_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if col_name in ['Negative', 'Negative %']:
            cell.fill = negative_fill
        elif col_name in ['Positive', 'Positive %']:
            cell.fill = positive_fill
        elif col_name in ['Neutral', 'Neutral %']:
            cell.fill = neutral_fill

for col in ws1.columns:
    max_len = max(len(str(cell.value)) for cell in col if cell.value)
    ws1.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 20)

ws1.freeze_panes = 'A2'

# Sheet 2: Overall Depot Performance
ws2 = wb.create_sheet("Overall Depot Performance")

for c_idx, col_name in enumerate(['Depot'] + overall_depot.columns.tolist(), 1):
    cell = ws2.cell(row=1, column=c_idx, value=col_name)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border

for r_idx, (depot, row_data) in enumerate(overall_depot.iterrows(), 2):
    ws2.cell(row=r_idx, column=1, value=depot).border = thin_border
    for col_idx, value in enumerate(row_data, 2):
        cell = ws2.cell(row=r_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

for col in ws2.columns:
    max_len = max(len(str(cell.value)) for cell in col if cell.value)
    ws2.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 20)

ws2.freeze_panes = 'A2'

# Sheet 3: Depot x Vehicle Correlation
print("\n" + "=" * 140)
print("DEPOT x VEHICLE CORRELATION")
print("=" * 140)
depot_vehicle = pd.crosstab(feedback_df['branch'].str.upper(), feedback_df['vehicle_type'], margins=True)
print(depot_vehicle)

ws3 = wb.create_sheet("Depot x Vehicle")
for col_idx, col_name in enumerate(depot_vehicle.columns, 1):
    cell = ws3.cell(row=1, column=col_idx, value=col_name)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border

for row_idx, (depot, row_data) in enumerate(depot_vehicle.iterrows(), 2):
    ws3.cell(row=row_idx, column=1, value=depot).border = thin_border
    for col_idx, value in enumerate(row_data, 2):
        cell = ws3.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

for col in ws3.columns:
    max_len = max(len(str(cell.value)) for cell in col if cell.value)
    ws3.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 15)

# Sheet 4: Complaint Categories (existing)
ws4 = wb.create_sheet("Complaint Categories")
complaint_dist = negative_feedback['complaint_category'].value_counts()
complaint_data = []
for cat in complaint_dist.index:
    count = complaint_dist[cat]
    pct = round(count / len(negative_feedback) * 100, 1)
    complaint_data.append({'Category': cat, 'Count': count, 'Percentage': pct})

complaint_df = pd.DataFrame(complaint_data)
for c_idx, col_name in enumerate(complaint_df.columns, 1):
    cell = ws4.cell(row=1, column=c_idx, value=col_name)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border

for r_idx, (_, row) in enumerate(complaint_df.iterrows(), 2):
    for c_idx, (col_name, value) in enumerate(row.items(), 1):
        cell = ws4.cell(row=r_idx, column=c_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

for col in ws4.columns:
    max_len = max(len(str(cell.value)) for cell in col if cell.value)
    ws4.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 20)

ws4.freeze_panes = 'A2'

wb.save('data_exports/Monthly_Feedback_Enhanced.xlsx')
print(f"\n✅ File saved: data_exports/Monthly_Feedback_Enhanced.xlsx")
print("\n📊 Sheets created:")
print("   1. Depot Monthly Breakdown - Ratings by depot per month")
print("   2. Overall Depot Performance - Best/worst performing depots")
print("   3. Depot x Vehicle - Vehicle distribution by depot")
print("   4. Complaint Categories - Distribution of complaint types")
