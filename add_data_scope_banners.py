"""
Add prominent data scope banners to export files to clarify that they show
feedback-linked deliveries only, and non-feedback-linked deliveries are excluded.
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

print("\n" + "=" * 100)
print("TASK: Add Data Scope Banners to Export Files")
print("=" * 100)

# ============================================================================
# FILE 1: Monthly_Feedback_Enhanced.xlsx
# ============================================================================
print("\n📄 Processing: Monthly_Feedback_Enhanced.xlsx")

file_path = 'data_exports/Monthly_Feedback_Enhanced.xlsx'
wb = load_workbook(file_path)

for sheet_name in wb.sheetnames:
    print(f"   Sheet: {sheet_name}")
    ws = wb[sheet_name]
    
    # Insert 3 rows at the top for banner
    ws.insert_rows(1, 3)
    
    # Row 1: Main banner (red background, white text, bold)
    banner_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    banner_font = Font(bold=True, color="FFFFFF", size=11)
    banner_border = Border(
        left=Side(style='medium'), 
        right=Side(style='medium'), 
        top=Side(style='medium'), 
        bottom=Side(style='medium')
    )
    
    main_cell = ws.merge_cells(f'A1:{get_column_letter(ws.max_column)}1')
    cell = ws['A1']
    cell.value = "⚠️ DATA SCOPE: This file contains FEEDBACK-LINKED DELIVERIES ONLY"
    cell.fill = banner_fill
    cell.font = banner_font
    cell.border = banner_border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 25
    
    # Row 2: Explanation (light red background)
    explanation_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
    explanation_font = Font(size=10)
    explanation_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    explanation_cell = ws.merge_cells(f'A2:{get_column_letter(ws.max_column)}2')
    cell = ws['A2']
    cell.value = "Non-feedback-linked deliveries (deliveries without feedback ratings) are EXCLUDED from this analysis"
    cell.fill = explanation_fill
    cell.font = explanation_font
    cell.border = explanation_border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[2].height = 20
    
    # Row 3: Coverage note
    coverage_fill = PatternFill(start_color="F9E6E6", end_color="F9E6E6", fill_type="solid")
    coverage_font = Font(size=9, italic=True)
    
    coverage_cell = ws.merge_cells(f'A3:{get_column_letter(ws.max_column)}3')
    cell = ws['A3']
    cell.value = "Coverage: 53,007 deliveries with feedback (36.5% of 150,000 total deliveries) | Analysis Period: Nov 2025 - May 2026"
    cell.fill = coverage_fill
    cell.font = coverage_font
    cell.border = explanation_border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[3].height = 18

wb.save(file_path)
print(f"   ✅ Saved with banner")

# ============================================================================
# FILE 2: Monthly_Feedback_Summary_Detailed.xlsx
# ============================================================================
print("\n📄 Processing: Monthly_Feedback_Summary_Detailed.xlsx")

file_path = 'data_exports/Monthly_Feedback_Summary_Detailed.xlsx'
wb = load_workbook(file_path)

for sheet_name in wb.sheetnames:
    print(f"   Sheet: {sheet_name}")
    ws = wb[sheet_name]
    
    # Insert 3 rows at the top
    ws.insert_rows(1, 3)
    
    # Row 1: Main banner
    banner_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    banner_font = Font(bold=True, color="FFFFFF", size=11)
    banner_border = Border(
        left=Side(style='medium'), 
        right=Side(style='medium'), 
        top=Side(style='medium'), 
        bottom=Side(style='medium')
    )
    
    main_cell = ws.merge_cells(f'A1:{get_column_letter(ws.max_column)}1')
    cell = ws['A1']
    cell.value = "⚠️ DATA SCOPE: This file contains FEEDBACK-LINKED DELIVERIES ONLY"
    cell.fill = banner_fill
    cell.font = banner_font
    cell.border = banner_border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 25
    
    # Row 2: Explanation
    explanation_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
    explanation_font = Font(size=10)
    explanation_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    explanation_cell = ws.merge_cells(f'A2:{get_column_letter(ws.max_column)}2')
    cell = ws['A2']
    cell.value = "Non-feedback-linked deliveries (deliveries without feedback ratings) are EXCLUDED from this analysis"
    cell.fill = explanation_fill
    cell.font = explanation_font
    cell.border = explanation_border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[2].height = 20
    
    # Row 3: Coverage note
    coverage_fill = PatternFill(start_color="F9E6E6", end_color="F9E6E6", fill_type="solid")
    coverage_font = Font(size=9, italic=True)
    
    coverage_cell = ws.merge_cells(f'A3:{get_column_letter(ws.max_column)}3')
    cell = ws['A3']
    cell.value = "Coverage: 53,007 feedback records analyzed (36.5% of 150,000 total deliveries) | Analysis Period: Nov 2025 - May 2026"
    cell.fill = coverage_fill
    cell.font = coverage_font
    cell.border = explanation_border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[3].height = 18

wb.save(file_path)
print(f"   ✅ Saved with banners (all {len(wb.sheetnames)} sheets)")

# ============================================================================
# FILE 3: deliveries_feedback_merged.csv
# ============================================================================
print("\n📄 Processing: deliveries_feedback_merged.csv")

# Read the CSV
df = pd.read_csv('data_exports/deliveries_feedback_merged.csv', encoding='latin-1')

# Create a new file with header comments
with open('data_exports/deliveries_feedback_merged.csv', 'w', encoding='utf-8') as f:
    # Write banner as comments
    f.write('# ====================================================================\n')
    f.write('# DATA SCOPE: This file contains FEEDBACK-LINKED DELIVERIES ONLY\n')
    f.write('# ====================================================================\n')
    f.write('# Non-feedback-linked deliveries (deliveries without feedback ratings)\n')
    f.write('# are EXCLUDED from this analysis.\n')
    f.write('#\n')
    f.write('# Coverage: 53,007 deliveries with feedback (36.5% of 150,000 total)\n')
    f.write('# Analysis Period: Nov 2025 - May 2026\n')
    f.write('# ====================================================================\n')
    f.write('\n')

# Append the CSV data
df.to_csv('data_exports/deliveries_feedback_merged.csv', mode='a', index=False, encoding='utf-8')
print(f"   ✅ Saved with header comments")

# ============================================================================
# FILE 4 & 5: priority_analysis.xlsx and vehicle_analysis.xlsx
# ============================================================================
for filename in ['priority_analysis.xlsx', 'vehicle_analysis.xlsx']:
    print(f"\n📄 Processing: {filename}")
    
    file_path = f'data_exports/{filename}'
    wb = load_workbook(file_path)
    
    for sheet_name in wb.sheetnames:
        print(f"   Sheet: {sheet_name}")
        ws = wb[sheet_name]
        
        # Insert 3 rows at the top
        ws.insert_rows(1, 3)
        
        # Row 1: Main banner
        banner_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        banner_font = Font(bold=True, color="FFFFFF", size=11)
        banner_border = Border(
            left=Side(style='medium'), 
            right=Side(style='medium'), 
            top=Side(style='medium'), 
            bottom=Side(style='medium')
        )
        
        main_cell = ws.merge_cells(f'A1:{get_column_letter(ws.max_column)}1')
        cell = ws['A1']
        cell.value = "⚠️ DATA SCOPE: This file contains FEEDBACK-LINKED DELIVERIES ONLY"
        cell.fill = banner_fill
        cell.font = banner_font
        cell.border = banner_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[1].height = 25
        
        # Row 2: Explanation
        explanation_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
        explanation_font = Font(size=10)
        explanation_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        explanation_cell = ws.merge_cells(f'A2:{get_column_letter(ws.max_column)}2')
        cell = ws['A2']
        cell.value = "Non-feedback-linked deliveries (deliveries without feedback ratings) are EXCLUDED from this analysis"
        cell.fill = explanation_fill
        cell.font = explanation_font
        cell.border = explanation_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[2].height = 20
        
        # Row 3: Coverage note
        coverage_fill = PatternFill(start_color="F9E6E6", end_color="F9E6E6", fill_type="solid")
        coverage_font = Font(size=9, italic=True)
        
        coverage_cell = ws.merge_cells(f'A3:{get_column_letter(ws.max_column)}3')
        cell = ws['A3']
        cell.value = "Coverage: 53,007 deliveries with feedback (36.5% of 150,000 total deliveries) | Analysis Period: Nov 2025 - May 2026"
        cell.fill = coverage_fill
        cell.font = coverage_font
        cell.border = explanation_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[3].height = 18
    
    wb.save(file_path)
    print(f"   ✅ Saved with banner")

print("\n" + "=" * 100)
print("✅ ALL FILES UPDATED WITH DATA SCOPE BANNERS")
print("=" * 100)

print("""
Summary of Changes:
  ✅ Monthly_Feedback_Enhanced.xlsx - Banner added to all sheets
  ✅ Monthly_Feedback_Summary_Detailed.xlsx - Banner added to all sheets
  ✅ deliveries_feedback_merged.csv - Header comments added
  ✅ priority_analysis.xlsx - Banner added to all sheets
  ✅ vehicle_analysis.xlsx - Banner added to all sheets

Banner Details:
  • Prominent red banner with white text
  • Clear statement: "FEEDBACK-LINKED DELIVERIES ONLY"
  • Explanation of scope (non-feedback deliveries excluded)
  • Coverage metrics (53,007 of 150,000 deliveries)
  • Analysis period (Nov 2025 - May 2026)

Result: Files now clearly communicate their data scope to any user or assessor.
""")
