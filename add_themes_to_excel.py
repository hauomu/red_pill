import pandas as pd
import sqlite3
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

print("📖 Reading existing Monthly_Feedback_Summary_Detailed.xlsx...")
existing_df = pd.read_excel('data_exports/Monthly_Feedback_Summary_Detailed.xlsx', sheet_name='Monthly Feedback Summary')
print(f"Loaded: {len(existing_df)} rows\n")

# Load feedback with themes
print("📊 Loading feedback data from database...")
conn = sqlite3.connect('delivery.db')

query = """
SELECT 
    strftime('%Y-%m', f.feedback_datetime) as Month,
    f.feedback_id,
    f.rating,
    f.comment
FROM feedback f
WHERE f.comment IS NOT NULL AND TRIM(f.comment) != ''
ORDER BY f.feedback_datetime
"""

feedback_df = pd.read_sql_query(query, conn)
conn.close()

print(f"Loaded: {len(feedback_df)} feedback records\n")

# Define themes and keywords
themes_dict = {
    'Delivery Speed': ['fast', 'slow', 'quick', 'delay', 'late', 'speed', 'on time', 'early', 'urgent', 'wait', 'took long', 'scheduled', 'window', 'promise'],
    'Service Quality': ['service', 'professional', 'polite', 'rude', 'friendly', 'helpful', 'attitude', 'courteous', 'behavior', 'unprofessional', 'responsive'],
    'Packaging/Condition': ['package', 'damage', 'broken', 'intact', 'box', 'wrapping', 'careful', 'careless', 'condition', 'packed', 'damaged', 'dent', 'crushed', 'looked roughly'],
    'Communication': ['communication', 'update', 'notification', 'informed', 'contacted', 'message', 'call', 'response', 'call ahead'],
    'Tracking': ['location', 'track', 'gps', 'address', 'wrong', 'lost', 'route', 'where'],
    'Driver Performance': ['driver', 'rider', 'delivery person', 'staff', 'excellent', 'poor', 'behavior'],
    'Cost/Pricing': ['price', 'cost', 'expensive', 'cheap', 'value', 'fee', 'charge', 'affordable'],
    'Other': []
}

def categorize_comment(comment):
    comment_lower = str(comment).lower() if comment else ""
    for theme, keywords in themes_dict.items():
        if theme == 'Other':
            continue
        for keyword in keywords:
            if keyword.lower() in comment_lower:
                return theme
    return 'Other'

feedback_df['theme'] = feedback_df['comment'].apply(categorize_comment)

# Group by Month and Theme to get counts and ratings
print("📊 Analyzing themes by month...")
monthly_themes = feedback_df.groupby(['Month', 'theme']).agg({
    'feedback_id': 'count',
    'rating': 'mean'
}).reset_index()
monthly_themes.columns = ['Month', 'Theme', 'Count', 'Avg_Rating']

print("\nMonthly Theme Breakdown:")
print(monthly_themes.head(20))

# Get top 5 positive and negative comments per theme overall
print("\n📝 Extracting top comments per theme...")
theme_comments = {}

for theme in feedback_df['theme'].unique():
    theme_df = feedback_df[feedback_df['theme'] == theme]
    
    positive = theme_df[theme_df['rating'] >= 4]['comment'].head(5).tolist()
    negative = theme_df[theme_df['rating'] <= 2]['comment'].head(5).tolist()
    
    theme_comments[theme] = {
        'count': len(theme_df),
        'pct': (len(theme_df) / len(feedback_df) * 100),
        'avg_rating': theme_df['rating'].mean(),
        'positive_comments': positive,
        'negative_comments': negative
    }

# Create new sheet with categories
print("\n📝 Creating categories sheet...")
wb = load_workbook('data_exports/Monthly_Feedback_Summary_Detailed.xlsx')

# Remove old categories sheet if exists
if 'Categories' in wb.sheetnames:
    del wb['Categories']

# Create new sheet
ws = wb.create_sheet('Categories', 1)

# Formatting
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                     top=Side(style='thin'), bottom=Side(style='thin'))
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)

# Title
ws['A1'] = "Feedback Categories & Theme Analysis"
ws['A1'].font = Font(bold=True, size=13, color="FFFFFF")
ws['A1'].fill = header_fill

# Headers
headers = ['Theme', 'Total Count', 'Percentage (%)', 'Avg Rating', 'Top Positive Comments', 'Top Negative Comments']
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=2, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Add theme data
row = 3
for theme in sorted(theme_comments.keys()):
    data = theme_comments[theme]
    
    ws.cell(row=row, column=1, value=theme)
    ws.cell(row=row, column=2, value=data['count'])
    ws.cell(row=row, column=3, value=round(data['pct'], 1))
    ws.cell(row=row, column=4, value=round(data['avg_rating'], 2))
    
    # Format positive comments
    positive_str = '\n'.join([f"{i}. {c[:60]}..." for i, c in enumerate(data['positive_comments'], 1)])
    ws.cell(row=row, column=5, value=positive_str)
    
    # Format negative comments
    negative_str = '\n'.join([f"{i}. {c[:60]}..." for i, c in enumerate(data['negative_comments'], 1)])
    ws.cell(row=row, column=6, value=negative_str)
    
    # Format row
    for col in range(1, 7):
        cell = ws.cell(row=row, column=col)
        cell.border = thin_border
        if col in [2, 3, 4]:
            cell.alignment = Alignment(horizontal='center', vertical='top')
            if col == 3:
                cell.number_format = '0.0'
            elif col == 4:
                cell.number_format = '0.00'
        else:
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    row += 1

# Auto-fit columns
ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 35
ws.column_dimensions['F'].width = 35

# Save
wb.save('data_exports/Monthly_Feedback_Summary_Detailed.xlsx')
print("✅ Saved: data_exports/Monthly_Feedback_Summary_Detailed.xlsx")

# Print summary
print("\n" + "=" * 100)
print("CATEGORY SUMMARY")
print("=" * 100)
for theme in sorted(theme_comments.keys()):
    data = theme_comments[theme]
    print(f"\n{theme}")
    print(f"  Count: {data['count']} | Percentage: {data['pct']:.1f}% | Avg Rating: {data['avg_rating']:.2f}")
    print(f"  Positive comments: {len(data['positive_comments'])}")
    print(f"  Negative comments: {len(data['negative_comments'])}")

print("\n" + "=" * 100)
print("✅ UPDATE COMPLETE")
print("=" * 100)
