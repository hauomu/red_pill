import openpyxl

# Check what sheets exist
try:
    wb = openpyxl.load_workbook('data_exports/Monthly_Feedback_Enhanced.xlsx')
    print("Sheets in Monthly_Feedback_Enhanced.xlsx:")
    for sheet in wb.sheetnames:
        print(f"  - {sheet}")
except Exception as e:
    print(f"Monthly_Feedback_Enhanced.xlsx not found: {e}")

try:
    wb = openpyxl.load_workbook('data_exports/Depot_Monthly_OnTime_Proportion.xlsx')
    print("\nSheets in Depot_Monthly_OnTime_Proportion.xlsx:")
    for sheet in wb.sheetnames:
        print(f"  - {sheet}")
except Exception as e:
    print(f"Depot_Monthly_OnTime_Proportion.xlsx: {e}")
