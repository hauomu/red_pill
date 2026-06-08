import pandas as pd
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

print("📊 Loading feedback data from database...")
conn = sqlite3.connect('delivery.db')

query = """
SELECT 
    f.feedback_id,
    f.rating,
    f.comment
FROM feedback f
WHERE f.comment IS NOT NULL AND TRIM(f.comment) != ''
ORDER BY f.feedback_id
"""

feedback_df = pd.read_sql_query(query, conn)
conn.close()

print(f"Loaded: {len(feedback_df)} feedback records\n")

# Define themes and keywords
themes_dict = {
    'Delivery Speed': ['fast', 'slow', 'quick', 'delay', 'late', 'speed', 'on time', 'early', 'urgent', 'wait', 'took long', 'scheduled', 'window', 'promise'],
    'Service Quality': ['service', 'professional', 'polite', 'rude', 'friendly', 'helpful', 'attitude', 'courteous', 'behavior', 'unprofessional', 'responsive'],
    'Packaging/Condition': ['package', 'damage', 'broken', 'intact', 'box', 'wrapping', 'careful', 'careless', 'condition', 'packed', 'damaged', 'dent', 'crushed', 'looked roughly'],
    'Communication': ['communication', 'update', 'notification', 'informed', 'contacted', 'message', 'call', 'response', 'call ahead'],
    'Tracking': ['location', 'track', 'gps', 'address', 'wrong', 'lost', 'route', 'where'],
    'Driver Performance': ['driver', 'rider', 'delivery person', 'staff', 'excellent', 'poor', 'behavior'],
    'Cost/Pricing': ['price', 'cost', 'expensive', 'cheap', 'value', 'fee', 'charge', 'affordable'],
    'Other': []
}

def categorize_comment(comment):
    comment_lower = str(comment).lower() if comment else ""
    for theme, keywords in themes_dict.items():
        if theme == 'Other':
            continue
        for keyword in keywords:
            if keyword.lower() in comment_lower:
                return theme
    return 'Other'

feedback_df['theme'] = feedback_df['comment'].apply(categorize_comment)

# Get top 5 positive and negative comments per theme
print("📝 Extracting top comments per theme...")
theme_comments = {}

for theme in sorted(feedback_df['theme'].unique()):
    theme_df = feedback_df[feedback_df['theme'] == theme]
    
    positive = theme_df[theme_df['rating'] >= 4]['comment'].head(5).tolist()
    negative = theme_df[theme_df['rating'] <= 2]['comment'].head(5).tolist()
    
    theme_comments[theme] = {
        'count': len(theme_df),
        'pct': (len(theme_df) / len(feedback_df) * 100),
        'avg_rating': theme_df['rating'].mean(),
        'positive_comments': positive,
        'negative_comments': negative
    }

# Create new workbook
print("📝 Creating new Categories worksheet...")
wb = Workbook()
ws = wb.active
ws.title = 'Categories'

# Formatting
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                     top=Side(style='thin'), bottom=Side(style='thin'))
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)

# Title
ws['A1'] = "Feedback Categories & Theme Analysis"
ws['A1'].font = Font(bold=True, size=13)
ws['A1'].fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
ws['A1'].font = Font(bold=True, size=13, color="FFFFFF")
ws.merge_cells('A1:F1')

# Headers
headers = ['Theme', 'Total Count', 'Percentage (%)', 'Avg Rating', 'Top 5 Positive Comments', 'Top 5 Negative Comments']
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=2, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Add theme data
row = 3
for theme in sorted(theme_comments.keys()):
    data = theme_comments[theme]
    
    ws.cell(row=row, column=1, value=theme).border = thin_border
    ws.cell(row=row, column=2, value=data['count']).border = thin_border
    ws.cell(row=row, column=3, value=round(data['pct'], 1)).border = thin_border
    ws.cell(row=row, column=4, value=round(data['avg_rating'], 2)).border = thin_border
    
    # Format positive comments
    positive_str = '\n'.join([f"{i}. {c[:55]}..." for i, c in enumerate(data['positive_comments'], 1)])
    positive_cell = ws.cell(row=row, column=5, value=positive_str if positive_str else "None")
    positive_cell.border = thin_border
    
    # Format negative comments
    negative_str = '\n'.join([f"{i}. {c[:55]}..." for i, c in enumerate(data['negative_comments'], 1)])
    negative_cell = ws.cell(row=row, column=6, value=negative_str if negative_str else "None")
    negative_cell.border = thin_border
    
    # Format numbers
    ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
    ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
    ws.cell(row=row, column=3).number_format = '0.0'
    ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')
    ws.cell(row=row, column=4).number_format = '0.00'
    
    # Text alignment
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='top')
    positive_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    negative_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    row += 1

# Auto-fit columns
ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 40
ws.column_dimensions['F'].width = 40

# Freeze headers
ws.freeze_panes = 'A3'

# Save
output_path = 'data_exports/Monthly_Feedback_Summary_With_Categories.xlsx'
wb.save(output_path)
print(f"✅ Saved: {output_path}\n")

# Print summary
print("=" * 100)
print("CATEGORY SUMMARY")
print("=" * 100)
for theme in sorted(theme_comments.keys()):
    data = theme_comments[theme]
    print(f"\n{theme}")
    print(f"  Count: {data['count']:,} | Percentage: {data['pct']:.1f}% | Avg Rating: {data['avg_rating']:.2f}")
    print(f"  Positive comments found: {len(data['positive_comments'])}")
    print(f"  Negative comments found: {len(data['negative_comments'])}")

print("\n" + "=" * 100)
print("✅ COMPLETE - File created successfully!")
print("=" * 100)
